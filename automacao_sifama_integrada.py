"""
Automação Integrada SIFAMA - ANTT
Sistema: ANTT - Agência Nacional de Transportes Terrestres

Funcionalidades:
- Consulta de Pagamento (Situação da Dívida)
- Inscrição de Autos na SERASA
"""

import time
import os
import random
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, ElementClickInterceptedException, ElementNotInteractableException, StaleElementReferenceException
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
import tkinter as tk
from tkinter import filedialog, messagebox, ttk, scrolledtext
from pathlib import Path
import threading

from logging_utils import Logger, _ts
from sifama_constantes_inscricao import (
    URL_LOGIN,
    XP_CAMPO_AUTO_CONSULTA,
    XP_LOGIN_BOTAO,
    XP_LOGIN_SENHA,
    XP_LOGIN_USUARIO,
    XP_POPUP_OK,
)


class BaseAutomacao:
    """Base comum das automações do SIFAMA."""
    def __init__(self, logger=None):
        self.driver = None
        self.wait = None
        self.logger = logger or Logger()
        self.pausado = False
        self.parar = False
        self.usuario_login = None  # Armazenar credenciais para refazer login em caso de erro
        self.senha_login = None
        
    def criar_driver(self, headless=False):
        """Cria o Chrome com as opções usadas no fluxo da automação."""
        try:
            if headless:
                self.logger.log("Iniciando navegador Chrome em modo headless (oculto)...", "INFO")
            else:
                self.logger.log("Iniciando navegador Chrome em modo anônimo...", "INFO")
            
            options = webdriver.ChromeOptions()
            options.add_argument('--incognito')
            options.add_argument('--disable-blink-features=AutomationControlled')
            options.add_experimental_option("excludeSwitches", ["enable-automation", "enable-logging"])
            options.add_experimental_option('useAutomationExtension', False)

            # Evita que o Chrome desacelere demais quando fica em segundo plano.
            options.add_argument('--disable-backgrounding-occluded-windows')
            options.add_argument('--disable-renderer-backgrounding')
            options.add_argument('--disable-background-timer-throttling')
            options.add_argument('--disable-background-networking')
            options.add_argument('--force-fieldtrials=BackgroundRendererProcessLimit/0')

            # Limpa parte do ruído que o Chrome costuma soltar.
            options.add_argument('--log-level=3')  # Suprime logs (0=INFO, 1=WARNING, 2=ERROR, 3=FATAL)
            options.add_argument('--disable-logging')
            options.add_argument('--disable-dev-shm-usage')
            options.add_argument('--silent')
            options.add_argument('--disable-gpu-logging')
            
            if headless:
                options.add_argument('--headless')
                options.add_argument('--disable-gpu')
            else:
                options.add_argument('--start-maximized')
            
            self.driver = webdriver.Chrome(options=options)
            self.wait = WebDriverWait(self.driver, 60)
            self.logger.log("Navegador iniciado com sucesso!", "SUCCESS")
            return True
        except Exception as e:
            self.logger.log(f"Erro ao iniciar navegador: {str(e)}", "ERROR")
            return False
    
    def fazer_login(self, usuario, senha):
        """Faz o login no SIFAMA."""
        try:
            # Guarda as credenciais para um eventual relogin.
            self.usuario_login = usuario
            self.senha_login = senha
            
            self.logger.log("Acessando página de login...", "INFO")
            self.driver.get(URL_LOGIN)
            time.sleep(2)
            
            campo_usuario = None
            campo_senha = None
            botao_entrar = None
            
            # Primeiro tenta pelo caminho mais direto.
            try:
                campo_usuario = self.wait.until(
                    EC.presence_of_element_located((By.XPATH, XP_LOGIN_USUARIO))
                )
                campo_senha = self.driver.find_element(By.XPATH, XP_LOGIN_SENHA)
                botao_entrar = self.driver.find_element(By.XPATH, XP_LOGIN_BOTAO)
            except:
                # Se falhar, tenta uma busca mais aberta pelos campos.
                try:
                    campos_input = self.driver.find_elements(By.TAG_NAME, "input")
                    for campo in campos_input:
                        campo_id = campo.get_attribute("id") or ""
                        campo_type = campo.get_attribute("type") or ""
                        if "usuario" in campo_id.lower() or "user" in campo_id.lower():
                            campo_usuario = campo
                        elif "senha" in campo_id.lower() or "password" in campo_id.lower() or campo_type == "password":
                            campo_senha = campo
                    
                    botoes = self.driver.find_elements(By.TAG_NAME, "input")
                    for botao in botoes:
                        botao_type = botao.get_attribute("type") or ""
                        botao_value = botao.get_attribute("value") or ""
                        if botao_type == "submit" or "entrar" in botao_value.lower() or "login" in botao_value.lower():
                            botao_entrar = botao
                            break
                except:
                    pass
            
            if not campo_usuario or not campo_senha or not botao_entrar:
                raise Exception("Não foi possível encontrar os campos de login.")
            
            self.logger.log("Preenchendo credenciais...", "INFO")
            campo_usuario.clear()
            campo_usuario.send_keys(usuario)
            campo_senha.clear()
            campo_senha.send_keys(senha)
            botao_entrar.click()
            
            time.sleep(3)
            self.logger.log("Login realizado com sucesso!", "SUCCESS")
            return True
            
        except Exception as e:
            self.logger.log(f"Erro ao fazer login: {str(e)}", "ERROR")
            return False
    
    def aguardar_pausa(self):
        """Segura a execução enquanto a automação estiver pausada."""
        while self.pausado and not self.parar:
            time.sleep(0.5)

    def _obter_texto_body(self, driver=None):
        """Lê o texto da página via JS para fugir de stale em postback."""
        d = driver or self.driver
        try:
            return d.execute_script("return document.body ? document.body.innerText : ''") or ""
        except Exception:
            return ""

    def _aguardar_overlay_sumir(self, timeout=8):
        """Espera os overlays sumirem antes de clicar para evitar clique interceptado."""
        try:
            for id_overlay in ("Progress_ModalProgress_backgroundElement", "Progress_UpdateProgress"):
                try:
                    el = self.driver.find_element(By.ID, id_overlay)
                    if el.is_displayed():
                        WebDriverWait(self.driver, timeout, poll_frequency=0.2).until(
                            EC.invisibility_of_element_located((By.ID, id_overlay))
                        )
                except NoSuchElementException:
                    pass
        except Exception:
            pass

    def verificar_erro_servidor(self):
        """Confere se a página caiu em erro de servidor."""
        try:
            # Aqui o texto via JS costuma ser mais confiável que pegar elemento.
            texto_pagina = self._obter_texto_body()
            if "Server Error" in texto_pagina or "Runtime Error" in texto_pagina or "Server Error in '/sar' Application" in texto_pagina:
                self.logger.log("Erro de servidor detectado na página! Tentando recarregar...", "WARNING")
                return True
            return False
        except:
            return False
    
    def tratar_erro_servidor(self, tentar_navegar_novamente=False):
        """Tenta se recuperar do erro abrindo nova guia e refazendo o caminho."""
        try:
            self.logger.log("=== INICIANDO TRATAMENTO DE ERRO DE SERVIDOR ===", "WARNING")
            self.logger.log("Abrindo nova guia e fechando a com erro...", "INFO")
            
            # Abre uma guia limpa.
            self.driver.execute_script("window.open('');")
            time.sleep(0.5)
            
            # Pega as guias abertas.
            janelas = self.driver.window_handles
            self.logger.log(f"Total de janelas encontradas: {len(janelas)}", "INFO")
            
            if len(janelas) < 2:
                self.logger.log("Erro: Não foi possível criar nova guia!", "ERROR")
                return False
            
            # Vai para a guia nova.
            self.driver.switch_to.window(janelas[-1])
            self.logger.log("Mudado para nova guia", "INFO")
            time.sleep(0.5)
            
            # Fecha a guia que ficou com erro.
            try:
                self.driver.switch_to.window(janelas[0])
                time.sleep(0.3)
                self.driver.close()
                self.logger.log("Aba antiga (com erro) fechada", "INFO")
                time.sleep(0.3)
            except Exception as e:
                self.logger.log(f"Aviso ao fechar aba antiga: {str(e)}", "WARNING")
            
            # Volta a trabalhar na guia boa.
            janelas_restantes = self.driver.window_handles
            if janelas_restantes:
                self.driver.switch_to.window(janelas_restantes[-1])
                self.logger.log("Voltado para nova guia", "INFO")
            else:
                self.logger.log("Erro: Nenhuma janela disponível após fechar aba antiga!", "ERROR")
                return False
            time.sleep(0.5)
            
            # Recomeça pelo login.
            self.logger.log("Navegando para página de login...", "INFO")
            self.driver.get(URL_LOGIN)
            time.sleep(2)
            
            # Se as credenciais estiverem guardadas, já tenta relogar.
            if self.usuario_login and self.senha_login:
                self.logger.log("Fazendo login novamente na nova guia...", "INFO")
                if not self.fazer_login(self.usuario_login, self.senha_login):
                    self.logger.log("Erro ao fazer login na nova guia!", "ERROR")
                    return False
                self.logger.log("Login realizado com sucesso na nova guia!", "SUCCESS")
                time.sleep(1)
            else:
                self.logger.log("Credenciais não disponíveis para refazer login.", "WARNING")
                return False
            
            # Se quem chamou quiser, a navegação completa pode ser refeita depois.
            if tentar_navegar_novamente:
                self.logger.log("=== TRATAMENTO CONCLUÍDO: Pronto para refazer navegação ===", "SUCCESS")
                return True
            
            # Se ainda sobrou algum erro, tenta refresh antes de devolver.
            if self.verificar_erro_servidor():
                self.logger.log("Erro de servidor ainda presente na nova guia. Aguardando...", "WARNING")
                time.sleep(3)
                self.driver.refresh()
                time.sleep(2)
            
            self.logger.log("=== TRATAMENTO CONCLUÍDO ===", "SUCCESS")
            return True
        except Exception as e:
            import traceback
            self.logger.log(f"Erro ao tratar erro de servidor: {str(e)}", "ERROR")
            self.logger.log(f"Traceback: {traceback.format_exc()}", "ERROR")
            # Último recurso: recarrega e devolve o controle.
            try:
                self.driver.refresh()
                time.sleep(2)
                return True
            except:
                return False

    @staticmethod
    def _mensagem_erro_navegacao(exc):
        """Monta uma mensagem útil quando o ChromeDriver volta quase sem detalhe."""
        msg = (str(exc) or "").strip()
        if msg:
            return msg
        msg = getattr(exc, "msg", None) or ""
        if msg and str(msg).strip():
            return str(msg).strip()
        return f"{type(exc).__name__} (sem mensagem do driver — possível timeout, overlay ou janela fechada)"

    def _hover_e_clicar_submenu(self, xpath_hover, xpath_link, descricao="menu", xpath_link_fallback=None):
        """
        Hover no item do menu principal e clique no link do submenu.
        Inclui espera de overlay, scroll, hover real + eventos JS de mouse,
        pausa para o submenu abrir, retry e clique via JS.
        """
        for tentativa in range(4):
            try:
                self._neutralizar_barra_governo()
                self._aguardar_overlay_sumir(timeout=10)
                hover_el = WebDriverWait(self.driver, 45, poll_frequency=0.3).until(
                    EC.presence_of_element_located((By.XPATH, xpath_hover))
                )
                self.driver.execute_script(
                    "arguments[0].scrollIntoView({block:'center', behavior:'instant'});", hover_el
                )
                time.sleep(0.10)
                ActionChains(self.driver).move_to_element(hover_el).pause(0.20).perform()
                self.driver.execute_script(
                    "var el=arguments[0];"
                    "['mouseover','mouseenter','mousemove'].forEach(function(evt){"
                    "  el.dispatchEvent(new MouseEvent(evt,{bubbles:true,cancelable:true,view:window}));"
                    "});",
                    hover_el,
                )
                time.sleep(0.45)
                self._aguardar_overlay_sumir(timeout=6)
                try:
                    link = WebDriverWait(self.driver, 22, poll_frequency=0.25).until(
                        EC.presence_of_element_located((By.XPATH, xpath_link))
                    )
                except TimeoutException:
                    if not xpath_link_fallback:
                        raise
                    link = WebDriverWait(self.driver, 8, poll_frequency=0.25).until(
                        EC.presence_of_element_located((By.XPATH, xpath_link_fallback))
                    )
                self.driver.execute_script(
                    "arguments[0].scrollIntoView({block:'center', behavior:'instant'});", link
                )
                self.driver.execute_script("arguments[0].click();", link)
                time.sleep(0.50)
                return True
            except Exception as exc:
                if xpath_link_fallback:
                    try:
                        clicou_fallback = self.driver.execute_script(
                            "var xp = arguments[0];"
                            "var el = document.evaluate(xp, document, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null).singleNodeValue;"
                            "if(!el) return false;"
                            "el.click();"
                            "return true;",
                            xpath_link_fallback,
                        )
                        if clicou_fallback:
                            time.sleep(0.50)
                            return True
                    except Exception:
                        pass
                self.logger.log(
                    f"Navegação «{descricao}» — tentativa {tentativa + 1}/4: {self._mensagem_erro_navegacao(exc)}",
                    "WARNING",
                )
                time.sleep(0.8 + tentativa * 0.25)
        return False

    def _neutralizar_barra_governo(self):
        """
        A barra amarela do governo (bgBarraAmarelaGoverno) pode interceptar cliques
        no topo durante transições/postbacks. Neutraliza ponteiros sem remover layout.
        """
        try:
            self.driver.execute_script(
                "var el=document.querySelector('.bgBarraAmarelaGoverno');"
                "if(el){el.style.pointerEvents='none';el.style.zIndex='0';}"
                "var el2=document.querySelector('.barraGoverno');"
                "if(el2){el2.style.pointerEvents='none';el2.style.zIndex='0';}"
            )
        except Exception:
            pass

    def _clicar_portal_sistemas_se_existir(self) -> bool:
        """Clica em Portal de Sistemas se o botão aparecer; retorna True se clicou, False se não existir."""
        xpath_portal = (
            '//*[@id="ContentPlaceHolderCorpo_ContentPlaceHolderCorpo_ContentPlaceHolderCorpo_'
            'ContentPlaceHolderCorpo_btnPortalSistemas"]'
        )
        try:
            self._aguardar_overlay_sumir(timeout=6)
            btn = WebDriverWait(self.driver, 6, poll_frequency=0.3).until(
                EC.element_to_be_clickable((By.XPATH, xpath_portal))
            )
            self.logger.log("Clicando no botão Portal de Sistemas...", "INFO")
            self.driver.execute_script("arguments[0].click();", btn)
            time.sleep(2.2)
            return True
        except TimeoutException:
            self.logger.log("Botão Portal de Sistemas não encontrado, continuando...", "INFO")
            return False
        except Exception as exc:
            self.logger.log(f"Portal de Sistemas: {self._mensagem_erro_navegacao(exc)} — continuando...", "WARNING")
            return False

    def fechar(self):
        """Fecha o navegador"""
        if self.driver:
            try:
                self.driver.quit()
                self.logger.log("Navegador fechado.", "INFO")
            except:
                pass


# ─────────────────────────────────────────────────────────────
# Checkpoint — retomar de onde parou
# ─────────────────────────────────────────────────────────────

import json as _json

class CheckpointManager:
    """Salva e carrega o progresso da automação para permitir retomar de onde parou."""

    def __init__(self, caminho_planilha: str):
        from pathlib import Path as _Path
        base = _Path(__file__).resolve().parent / "checkpoints"
        base.mkdir(exist_ok=True)
        nome = _Path(caminho_planilha).stem.replace(" ", "_")
        self.path = base / f"checkpoint_{nome}.json"

    def existe(self) -> bool:
        return self.path.exists()

    def salvar(self, idx_atual: int, total: int, resultados: list):
        try:
            with open(self.path, "w", encoding="utf-8") as f:
                _json.dump({"idx": idx_atual, "total": total, "resultados": resultados}, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    def carregar(self) -> dict:
        try:
            with open(self.path, "r", encoding="utf-8") as f:
                return _json.load(f)
        except Exception:
            return {}

    def limpar(self):
        try:
            if self.path.exists():
                self.path.unlink()
        except Exception:
            pass


class AutomacaoConsultaPagamento(BaseAutomacao):
    """Fluxo de consulta de pagamento e situação da dívida."""
    
    def __init__(self, logger=None):
        super().__init__(logger)
        self.resultados = []
    
    def navegar_para_formulario(self):
        """Navega até a tela de consulta."""
        xp_menu0 = '//*[@id="ContentPlaceHolderCorpo_ContentPlaceHolderMenu_MenuSistemasn0"]/table/tbody/tr/td/a'
        xp_menu5 = '//*[@id="ContentPlaceHolderCorpo_ContentPlaceHolderMenu_MenuSistemasn5"]/td/table/tbody/tr/td/a'
        xp_menu5_abs = '/html/body/form/div[4]/div[3]/table/tbody/tr/td[1]/div[8]/table/tbody/tr[1]/td/table/tbody/tr/td/a'
        xp_m2 = '//*[@id="ContentPlaceHolderCorpo_ContentPlaceHolderMenu_menun2"]/table/tbody/tr/td/a'
        xp_m19 = '//*[@id="ContentPlaceHolderCorpo_ContentPlaceHolderMenu_menun19"]/td/table/tbody/tr/td/a'
        xp_campo = XP_CAMPO_AUTO_CONSULTA
        try:
            self.logger.log("Navegando até o formulário de consulta...", "INFO")
            self._aguardar_overlay_sumir(timeout=12)
            if not self._hover_e_clicar_submenu(xp_menu0, xp_menu5, "Sistemas → submenu", xp_menu5_abs):
                raise TimeoutException("Falha ao abrir primeiro submenu do menu Sistemas")
            if self._clicar_portal_sistemas_se_existir():
                self.logger.log("Clicando novamente nos botões de navegação...", "INFO")
                if not self._hover_e_clicar_submenu(xp_menu0, xp_menu5, "Sistemas → submenu (após Portal)", xp_menu5_abs):
                    raise TimeoutException("Falha ao repetir menu após Portal de Sistemas")
            if not self._hover_e_clicar_submenu(xp_m2, xp_m19, "Consulta → formulário"):
                raise TimeoutException("Falha ao abrir formulário de consulta (menu nível 2)")
            WebDriverWait(self.driver, 40, poll_frequency=0.3).until(
                EC.presence_of_element_located((By.XPATH, xp_campo))
            )
            if self.verificar_erro_servidor():
                self.logger.log("Erro de servidor detectado após navegação!", "ERROR")
                if not self.tratar_erro_servidor(tentar_navegar_novamente=False):
                    return False
            self.logger.log("Formulário de consulta carregado!", "SUCCESS")
            return True
        except Exception as e:
            self.logger.log(f"Erro na navegação: {self._mensagem_erro_navegacao(e)}", "ERROR")
            return False
    
    def consultar_auto(self, numero_auto):
        """Consulta um auto e devolve o resultado da pesquisa."""
        try:
            self.aguardar_pausa()
            if self.parar:
                return (False, False)
            
            self.logger.log(f"Consultando auto: {numero_auto}...", "INFO")
            
            # Preenche o auto.
            campo_auto = self.wait.until(
                EC.presence_of_element_located((By.XPATH, XP_CAMPO_AUTO_CONSULTA))
            )
            campo_auto.clear()
            campo_auto.send_keys(str(numero_auto))
            
            # Dispara a consulta.
            botao_gerar = self.wait.until(
                EC.element_to_be_clickable((By.XPATH, '//*[@id="ContentPlaceHolderCorpo_ContentPlaceHolderCorpo_ContentPlaceHolderCorpo_ContentPlaceHolderCorpo_btnGerar"]'))
            )
            botao_gerar.click()
            
            # Dá um tempo curto para o portal processar.
            time.sleep(2)
            
            # Se o portal caiu em erro, tenta se recuperar antes de desistir.
            if self.verificar_erro_servidor():
                self.logger.log("Erro de servidor detectado após clicar em Gerar!", "ERROR")
                if self.tratar_erro_servidor(tentar_navegar_novamente=True):
                    # Tentar refazer navegação
                    if not self.navegar_para_formulario():
                        self.logger.log("Não foi possível navegar até o formulário após tratar erro!", "ERROR")
                        return (False, False)
                    # Tentar consultar novamente o MESMO auto
                    self.logger.log(f"Retentando consulta do auto {numero_auto} após tratar erro...", "INFO")
                    campo_auto = self.wait.until(
                        EC.presence_of_element_located((By.XPATH, XP_CAMPO_AUTO_CONSULTA))
                    )
                    campo_auto.clear()
                    campo_auto.send_keys(str(numero_auto))
                    botao_gerar = self.wait.until(
                        EC.element_to_be_clickable((By.XPATH, '//*[@id="ContentPlaceHolderCorpo_ContentPlaceHolderCorpo_ContentPlaceHolderCorpo_ContentPlaceHolderCorpo_btnGerar"]'))
                    )
                    botao_gerar.click()
                    time.sleep(2)
                    # Verificar novamente se ainda tem erro
                    if self.verificar_erro_servidor():
                        self.logger.log("Erro de servidor persiste após retentar!", "ERROR")
                        return (False, False)
                else:
                    self.logger.log("Falha ao tratar erro de servidor!", "ERROR")
                    return (False, False)
            
            # Verificar se aparece o popup "Nenhum registro encontrado" e clicar em Ok
            popup_encontrado = False
            try:
                # Usar timeout curto para não atrasar quando não houver popup
                botao_ok = WebDriverWait(self.driver, 3).until(
                    EC.element_to_be_clickable((By.XPATH, XP_POPUP_OK))
                )
                self.logger.log("Popup de 'Nenhum registro encontrado' detectado, clicando em Ok...", "INFO")
                botao_ok.click()
                popup_encontrado = True
                time.sleep(0.5)
            except:
                # Se não aparecer o popup, continua normalmente
                pass
            
            time.sleep(0.5)
            # Retornar tupla: (sucesso, popup_encontrado)
            return (True, popup_encontrado)
            
        except Exception as e:
            self.logger.log(f"Erro ao consultar auto {numero_auto}: {str(e)}", "ERROR")
            return (False, False)
    
    def extrair_situacao_divida(self):
        """Extrai a situação da dívida usando os caminhos mais estáveis que já deram certo."""
        try:
            time.sleep(1)  # Reduzido de 2s para 1s
            
            # Antes de extrair, confere se a tela ainda está íntegra.
            if self.verificar_erro_servidor():
                self.logger.log("Erro de servidor detectado ao tentar extrair situação!", "ERROR")
                return None
            
            # O cuidado aqui é não pegar o label no lugar do valor real.
            
            # Começa pelos casos mais diretos.
            try:
                # Buscar div que contém "Quitada" ou "Quitado"
                try:
                    div_quitada = self.driver.find_element(
                        By.XPATH,
                        "//div[contains(@style, 'width:15.21mm') and contains(@style, 'min-width: 15.21mm') and (contains(text(), 'Quitada') or contains(text(), 'Quitado'))]"
                    )
                    texto = div_quitada.text.strip()
                    if texto and texto.lower() in ['quitada', 'quitado']:
                        self.logger.log(f"Situação encontrada: {texto}", "SUCCESS")
                        return texto
                except:
                    pass
                
                # Buscar div que contém "Pendente"
                try:
                    div_pendente = self.driver.find_element(
                        By.XPATH,
                        "//div[contains(@style, 'width:15.21mm') and contains(@style, 'min-width: 15.21mm') and contains(text(), 'Pendente')]"
                    )
                    texto = div_pendente.text.strip()
                    if texto and texto.lower() == 'pendente':
                        self.logger.log(f"Situação encontrada: {texto}", "SUCCESS")
                        return texto
                except:
                    pass
                
            except Exception as e:
                self.logger.log(f"Erro na estratégia 1: {str(e)}", "WARNING")
            
            # Se não achou de primeira, faz uma varredura mais ampla.
            try:
                divs = self.driver.find_elements(
                    By.XPATH, 
                    "//div[contains(@style, 'width:15.21mm') and contains(@style, 'min-width: 15.21mm')]"
                )
                
                self.logger.log(f"Encontrados {len(divs)} divs com o estilo específico", "INFO")
                
                # Filtrar divs que NÃO são o label "Situação da Dívida"
                for div in divs:
                    texto = div.text.strip()
                    if not texto:
                        continue
                    
                    texto_lower = texto.lower()
                    
                    # IGNORAR COMPLETAMENTE se contiver "Situação" e "Dívida" ou "Divida"
                    if ('situação' in texto_lower or 'situacao' in texto_lower) and ('divida' in texto_lower or 'dívida' in texto_lower):
                        self.logger.log(f"Ignorando label/cabeçalho: '{texto}'", "INFO")
                        continue
                    
                    # ACEITAR apenas se for "Quitada", "Pendente" ou "Quitado"
                    if texto_lower in ['quitada', 'pendente', 'quitado']:
                        self.logger.log(f"Situação encontrada: {texto}", "SUCCESS")
                        return texto
                    # Se for um texto muito curto (menos de 15 caracteres) e não for o label, pode ser válido
                    elif len(texto) < 15 and texto_lower not in ['situação', 'situacao', 'divida', 'dívida', 'situação da dívida', 'situacao da divida']:
                        # Verificar se não contém palavras do label
                        palavras_label = ['situação', 'situacao', 'divida', 'dívida', 'da']
                        if not any(palavra in texto_lower for palavra in palavras_label):
                            self.logger.log(f"Situação encontrada (texto curto válido): {texto}", "SUCCESS")
                            return texto
                    else:
                        # Logar para debug mas NÃO retornar
                        self.logger.log(f"Div ignorado com texto: '{texto}' (não é valor válido)", "WARNING")
                
                self.logger.log("Nenhum div com valor válido encontrado na estratégia 2", "WARNING")
                
            except Exception as e:
                self.logger.log(f"Erro na estratégia 2: {str(e)}", "WARNING")
            
            # Outra tentativa excluindo o label logo no XPath.
            try:
                # Buscar div que NÃO contém "Situação" e "Dívida" mas contém o estilo
                divs_validos = self.driver.find_elements(
                    By.XPATH,
                    "//div[contains(@style, 'width:15.21mm') and contains(@style, 'min-width: 15.21mm') and normalize-space(text()) != '' and not(contains(text(), 'Situação') and contains(text(), 'Dívida')) and not(contains(text(), 'Situacao') and contains(text(), 'Divida'))]"
                )
                
                for div in divs_validos:
                    texto = div.text.strip()
                    texto_lower = texto.lower()
                    
                    # Aceitar apenas valores válidos
                    if texto_lower in ['quitada', 'pendente', 'quitado']:
                        self.logger.log(f"Situação encontrada (estratégia 3): {texto}", "SUCCESS")
                        return texto
                    elif len(texto) < 15:
                        # Texto curto que não é o label
                        palavras_label = ['situação', 'situacao', 'divida', 'dívida']
                        if not any(palavra in texto_lower for palavra in palavras_label):
                            self.logger.log(f"Situação encontrada (estratégia 3 - texto curto): {texto}", "SUCCESS")
                            return texto
                            
            except Exception as e:
                self.logger.log(f"Estratégia 3 falhou: {str(e)}", "WARNING")
            
            # Último caminho: achar o texto e subir para o elemento útil.
            try:
                texto_pagina = self._obter_texto_body()
                
                if "Quitada" in texto_pagina or "Quitado" in texto_pagina:
                    # Buscar o div que contém "Quitada" e tem o estilo
                    try:
                        div_quitada = self.driver.find_element(
                            By.XPATH,
                            "//div[contains(@style, 'width:15.21mm') and (contains(text(), 'Quitada') or contains(text(), 'Quitado')) and not(contains(text(), 'Situação') and contains(text(), 'Dívida'))]"
                        )
                        situacao = div_quitada.text.strip()
                        if situacao and situacao.lower() in ['quitada', 'quitado']:
                            self.logger.log(f"Situação encontrada (estratégia 4 - Quitada): {situacao}", "SUCCESS")
                            return situacao
                    except:
                        pass
                        
                if "Pendente" in texto_pagina:
                    # Buscar o div que contém "Pendente" e tem o estilo
                    try:
                        div_pendente = self.driver.find_element(
                            By.XPATH,
                            "//div[contains(@style, 'width:15.21mm') and contains(text(), 'Pendente') and not(contains(text(), 'Situação') and contains(text(), 'Dívida'))]"
                        )
                        situacao = div_pendente.text.strip()
                        if situacao and situacao.lower() == 'pendente':
                            self.logger.log(f"Situação encontrada (estratégia 4 - Pendente): {situacao}", "SUCCESS")
                            return situacao
                    except:
                        pass
                        
            except Exception as e:
                self.logger.log(f"Estratégia 4 falhou: {str(e)}", "WARNING")
            
            # Melhor devolver vazio do que devolver o texto errado.
            self.logger.log("Não foi possível extrair a situação da dívida (valor real não encontrado)!", "ERROR")
            return None
            
        except Exception as e:
            self.logger.log(f"Erro ao extrair situação: {str(e)}", "ERROR")
            import traceback
            self.logger.log(f"Traceback: {traceback.format_exc()}", "ERROR")
            return None
    
    def processar_autos(self, autos, progress_callback=None, stats_callback=None, error_handler=None):
        """Processa todos os autos"""
        total = len(autos)
        sucessos = 0
        erros = 0
        
        for idx, auto in enumerate(autos, 1):
            if self.parar:
                break
            
            auto = str(auto).strip()
            if not auto:
                continue
            
            try:
                if progress_callback:
                    progress_callback(f"Processando {idx}/{total}: {auto}")
                
                # Consultar auto
                resultado_consulta = self.consultar_auto(auto)
                # resultado_consulta é sempre uma tupla: (sucesso, popup_encontrado)
                sucesso_consulta, popup_encontrado = resultado_consulta
                
                if not sucesso_consulta:
                    self.resultados.append({'auto': auto, 'situacao': 'erro'})
                    erros += 1
                    if stats_callback:
                        stats_callback(sucessos, erros)
                    # Tentar refazer navegação se falhou
                    try:
                        if not self.navegar_para_formulario():
                            self.logger.log("Não foi possível navegar até o formulário. Interrompendo...", "ERROR")
                            break
                    except:
                        pass
                    continue
                
                # Verificar se popup foi encontrado (auto não encontrado)
                
                if popup_encontrado:
                    # Auto não encontrado no sistema
                    self.resultados.append({'auto': auto, 'situacao': 'NÃO ENCONTRADO'})
                    erros += 1
                    if stats_callback:
                        stats_callback(sucessos, erros)
                    continue
                
                # Verificar se há erro de servidor antes de extrair
                if self.verificar_erro_servidor():
                    self.logger.log(f"Auto {auto}: Erro de servidor detectado antes de extrair!", "ERROR")
                    # Tentar tratar erro (abrir nova guia, refazer login e navegação)
                    if self.tratar_erro_servidor(tentar_navegar_novamente=True):
                        if not self.navegar_para_formulario():
                            self.logger.log("Não foi possível navegar até o formulário após tratar erro. Interrompendo...", "ERROR")
                            break
                        # Retentar consulta do mesmo auto
                        self.logger.log(f"Retentando consulta do auto {auto} após tratar erro...", "INFO")
                        resultado_consulta = self.consultar_auto(auto)
                        if not resultado_consulta[0]:
                            self.logger.log(f"Auto {auto}: Falha ao consultar após tratar erro!", "ERROR")
                            self.resultados.append({'auto': auto, 'situacao': 'ERRO DE SERVIDOR'})
                            erros += 1
                            if stats_callback:
                                stats_callback(sucessos, erros)
                            continue
                        # Se consultou com sucesso, continuar para extrair situação
                    else:
                        self.logger.log(f"Auto {auto}: Falha ao tratar erro de servidor!", "ERROR")
                        self.resultados.append({'auto': auto, 'situacao': 'ERRO DE SERVIDOR'})
                        erros += 1
                        if stats_callback:
                            stats_callback(sucessos, erros)
                        continue
                
                # Extrair situação
                max_tentativas = 10
                situacao = None
                for tentativa in range(max_tentativas):
                    time.sleep(2)
                    situacao = self.extrair_situacao_divida()
                    if situacao:
                        break
                    # Verificar erro de servidor durante tentativas
                    if self.verificar_erro_servidor():
                        self.logger.log(f"Auto {auto}: Erro de servidor detectado durante extração!", "ERROR")
                        # Tentar tratar erro e retentar extração
                        if self.tratar_erro_servidor(tentar_navegar_novamente=True):
                            if not self.navegar_para_formulario():
                                self.logger.log("Não foi possível navegar após tratar erro durante extração!", "ERROR")
                                situacao = 'ERRO DE SERVIDOR'
                                break
                            # Retentar consulta e extração do mesmo auto
                            self.logger.log(f"Retentando consulta e extração do auto {auto}...", "INFO")
                            resultado_consulta = self.consultar_auto(auto)
                            if resultado_consulta[0]:
                                # Aguardar um pouco e tentar extrair novamente
                                time.sleep(2)
                                situacao = self.extrair_situacao_divida()
                                if situacao:
                                    break
                            else:
                                situacao = 'ERRO DE SERVIDOR'
                                break
                        else:
                            situacao = 'ERRO DE SERVIDOR'
                            break
                
                if not situacao:
                    situacao = 'erro'
                    self.logger.log(f"Auto {auto}: Não foi possível extrair situação (retornou None)", "ERROR")
                else:
                    # Log detalhado do que foi extraído
                    self.logger.log(f"Auto {auto}: Situação extraída = '{situacao}' (tipo: {type(situacao).__name__})", "INFO")
                
                # Garantir que a situação seja uma string válida
                if situacao and isinstance(situacao, str):
                    situacao = situacao.strip()
                    # Se a situação for muito longa ou contiver "Situação", pode estar errado
                    if len(situacao) > 50 or "Situação" in situacao or "situação" in situacao.lower():
                        self.logger.log(f"Auto {auto}: Situação suspeita detectada: '{situacao}'. Tentando re-extrair...", "WARNING")
                        # Tentar extrair novamente
                        time.sleep(1)
                        situacao_nova = self.extrair_situacao_divida()
                        if situacao_nova and len(situacao_nova) < 50:
                            situacao = situacao_nova.strip()
                            self.logger.log(f"Auto {auto}: Situação corrigida para: '{situacao}'", "SUCCESS")
                
                self.resultados.append({'auto': auto, 'situacao': situacao})
                
                if situacao and situacao != 'erro' and situacao.lower() not in ['erro', 'erro de servidor']:
                    sucessos += 1
                    self.logger.log(f"Auto {auto}: {situacao}", "SUCCESS")
                else:
                    erros += 1
                    self.logger.log(f"Auto {auto}: Erro ao extrair situação (valor final: '{situacao}')", "ERROR")
                
                if stats_callback:
                    stats_callback(sucessos, erros)
                
                # Limpar campo
                try:
                    campo_auto = self.driver.find_element(By.XPATH, XP_CAMPO_AUTO_CONSULTA)
                    campo_auto.clear()
                except:
                    if not self.navegar_para_formulario():
                        break
                
            except Exception as e:
                self.logger.log(f"Erro ao processar auto {auto}: {str(e)}", "ERROR")
                self.resultados.append({'auto': auto, 'situacao': 'erro'})
                erros += 1
                if stats_callback:
                    stats_callback(sucessos, erros)
        
        return sucessos, erros
    
    def salvar_resultados(self, caminho_original):
        """Salva os resultados na planilha com cabeçalhos corretos e formatação"""
        try:
            caminho = Path(caminho_original)
            
            # Criar DataFrame com os resultados
            dados = []
            for resultado in self.resultados:
                situacao = resultado['situacao']
                # Log para debug
                self.logger.log(f"Salvando auto {resultado['auto']} com situação: '{situacao}'", "INFO")
                
                # Validar que a situação não seja "Situação Divida" ou similar
                if situacao and isinstance(situacao, str):
                    situacao_lower = situacao.lower().strip()
                    if 'situação' in situacao_lower or 'situacao' in situacao_lower:
                        if 'divida' in situacao_lower or 'dívida' in situacao_lower:
                            # Se for "Situação Divida", tentar encontrar o valor real
                            self.logger.log(f"AVISO: Situação parece ser o cabeçalho '{situacao}' ao invés do valor real!", "WARNING")
                            # Manter como está, mas logar o problema
                            situacao = 'ERRO: Valor não extraído corretamente'
                
                dados.append({
                    'Autos de infração': resultado['auto'],
                    'Situação Divida': situacao
                })
            
            df_resultado = pd.DataFrame(dados)
            self.logger.log(f"Total de {len(dados)} registros preparados para salvar", "INFO")
            
            # Se for CSV, criar novo CSV
            if caminho.suffix.lower() == '.csv':
                novo_caminho = caminho.parent / "Planilha Consulta Pagamento Resultado.csv"
                df_resultado.to_csv(novo_caminho, index=False, encoding='utf-8')
            else:
                # Se for Excel
                novo_caminho = caminho.parent / "Planilha Consulta Pagamento Resultado.xlsx"
                df_resultado.to_excel(novo_caminho, index=False, engine='openpyxl')
                
                # Aplicar formatação (tabela e cabeçalho colorido)
                self._aplicar_formatacao_excel(novo_caminho, 'Situação Divida')
            
            self.logger.log(f"Resultados salvos em: {novo_caminho}", "SUCCESS")
            return novo_caminho
            
        except Exception as e:
            self.logger.log(f"Erro ao salvar resultados: {str(e)}", "ERROR")
            return None
    
    def _aplicar_formatacao_excel(self, caminho_excel, nome_coluna_situacao):
        """Aplica formatação de tabela e cabeçalho colorido na planilha Excel"""
        try:
            wb = load_workbook(caminho_excel)
            ws = wb.active
            
            # Definir largura das colunas
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 25
            
            # Formatar cabeçalho (linha 1)
            fill_cabecalho = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            font_cabecalho = Font(bold=True, color="FFFFFF", size=11)
            alignment_cabecalho = Alignment(horizontal="center", vertical="center")
            
            for col in range(1, 3):  # Colunas A e B
                cell = ws.cell(row=1, column=col)
                cell.fill = fill_cabecalho
                cell.font = font_cabecalho
                cell.alignment = alignment_cabecalho
            
            # Adicionar bordas
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Aplicar bordas em todas as células com dados
            max_row = ws.max_row
            for row in range(1, max_row + 1):
                for col in range(1, 3):
                    ws.cell(row=row, column=col).border = thin_border
            
            # Criar tabela (formato de tabela Excel)
            if max_row > 1:
                tab = Table(displayName="TabelaResultados", ref=f"A1:B{max_row}")
                style = TableStyleInfo(
                    name="TableStyleMedium9",
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False
                )
                tab.tableStyleInfo = style
                ws.add_table(tab)
            
            # Centralizar dados nas células
            for row in range(2, max_row + 1):
                for col in range(1, 3):
                    ws.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center")
            
            wb.save(caminho_excel)
            
        except Exception as e:
            self.logger.log(f"Erro ao aplicar formatação: {str(e)}", "WARNING")


class AutomacaoInscricaoSerasa(BaseAutomacao):
    """Fluxo de pesquisa e seleção para inscrição na SERASA."""
    
    def __init__(self, logger=None):
        super().__init__(logger)
        self.resultados = []
    
    def navegar_para_formulario(self):
        """Navega até a tela de inscrição na SERASA."""
        xp_menu0 = '//*[@id="ContentPlaceHolderCorpo_ContentPlaceHolderMenu_MenuSistemasn0"]/table/tbody/tr/td/a'
        xp_menu5 = '//*[@id="ContentPlaceHolderCorpo_ContentPlaceHolderMenu_MenuSistemasn5"]/td/table/tbody/tr/td/a'
        xp_menu5_abs = '/html/body/form/div[4]/div[3]/table/tbody/tr/td[1]/div[8]/table/tbody/tr[1]/td/table/tbody/tr/td/a'
        xp_serasa = '//*[@id="ContentPlaceHolderCorpo_ContentPlaceHolderMenu_menun4"]/table/tbody/tr/td/a'
        xp_inscricao = '//*[@id="ContentPlaceHolderCorpo_ContentPlaceHolderMenu_menun43"]/td/table/tbody/tr/td/a'
        xp_campo = '//*[@id="Corpo_txbAutoInfracao"]'
        try:
            self.logger.log("Navegando até o formulário de inscrição SERASA...", "INFO")
            self._aguardar_overlay_sumir(timeout=12)
            if not self._hover_e_clicar_submenu(xp_menu0, xp_menu5, "Sistemas → submenu", xp_menu5_abs):
                raise TimeoutException("Falha ao abrir primeiro submenu do menu Sistemas")
            if self._clicar_portal_sistemas_se_existir():
                self.logger.log("Clicando novamente nos botões de navegação...", "INFO")
                if not self._hover_e_clicar_submenu(xp_menu0, xp_menu5, "Sistemas → submenu (após Portal)", xp_menu5_abs):
                    raise TimeoutException("Falha ao repetir menu após Portal de Sistemas")
            if not self._hover_e_clicar_submenu(xp_serasa, xp_inscricao, "Serasa → Inscrição"):
                raise TimeoutException("Falha ao abrir formulário de inscrição SERASA (menu Serasa)")
            WebDriverWait(self.driver, 40, poll_frequency=0.3).until(
                EC.presence_of_element_located((By.XPATH, xp_campo))
            )
            if self.verificar_erro_servidor():
                self.logger.log("Erro de servidor detectado após navegação!", "ERROR")
                if not self.tratar_erro_servidor(tentar_navegar_novamente=False):
                    return False
            self.logger.log("Formulário de inscrição SERASA carregado!", "SUCCESS")
            return True
        except Exception as e:
            self.logger.log(f"Erro na navegação: {self._mensagem_erro_navegacao(e)}", "ERROR")
            return False
    
    def pesquisar_auto(self, numero_auto):
        """Pesquisa um auto na tela da SERASA."""
        try:
            self.aguardar_pausa()
            if self.parar:
                return False

            self.logger.log(f"Pesquisando auto: {numero_auto}...", "INFO")

            # Preenche o campo e já esconde a barra que costuma atrapalhar.
            self.driver.execute_script(
                "var b=document.getElementById('wings_process_presentation_dashboard_bar');"
                "if(b)b.style.display='none';"
                "var f=document.getElementById('Corpo_txbAutoInfracao');"
                "if(f){f.value=arguments[0];f.dispatchEvent(new Event('input',{bubbles:true}));}",
                str(numero_auto)
            )

            tentativa = 0
            while not self.parar:
                tentativa += 1

                # Se o overlay ainda estiver visível, segura um pouco antes do clique.
                self._aguardar_overlay_sumir(timeout=8)
                # O clique por JS aqui costuma ser mais estável que manter a referência do botão.
                clicou = False
                for _t in range(3):
                    try:
                        clicou = self.driver.execute_script(
                            "var btn=document.getElementById('Corpo_btnPesquisar');"
                            "if(!btn)return false;"
                            "btn.click();"
                            "return true;"
                        )
                        if clicou:
                            break
                    except Exception:
                        pass
                    time.sleep(0.35)
                if not clicou:
                    self.logger.log(
                        f"Botão Pesquisar não encontrado ou falha ao clicar para {numero_auto}.", "WARNING"
                    )
                    return False

                # Espera resultado ou a mensagem de que não houve retorno.
                try:
                    def _checkbox_ou_nenhum(d):
                        try:
                            if d.find_elements(By.CSS_SELECTOR, "[id^='Corpo_gdvAutoInfracao_ckSelecionar_']"):
                                return True
                            if "Nenhum registro" in self._obter_texto_body(d):
                                return True
                        except StaleElementReferenceException:
                            pass  # DOM em atualização, continua polling
                        return False
                    WebDriverWait(self.driver, 10, poll_frequency=0.25).until(_checkbox_ou_nenhum)
                except TimeoutException:
                    pass

                # Verificar se há erro de servidor
                if self.verificar_erro_servidor():
                    self.logger.log("Erro de servidor detectado após clicar em Pesquisar!", "ERROR")
                    if self.tratar_erro_servidor(tentar_navegar_novamente=True):
                        if not self.navegar_para_formulario():
                            self.logger.log("Não foi possível navegar até o formulário após tratar erro!", "ERROR")
                            return False
                        self.logger.log(f"Retentando pesquisa do auto {numero_auto} após tratar erro de servidor...", "INFO")
                        self.driver.execute_script(
                            "var f=document.getElementById('Corpo_txbAutoInfracao');"
                            "if(f){f.value=arguments[0];f.dispatchEvent(new Event('input',{bubbles:true}));}",
                            str(numero_auto)
                        )
                        continue
                    else:
                        self.logger.log("Falha ao tratar erro de servidor!", "ERROR")
                        return False

                # Depois do postback, a grid ainda leva um instante para assentar.
                time.sleep(2.0)

                # Pesquisa concluída
                return True

            return False

        except StaleElementReferenceException:
            # Postback recriou o DOM durante a pesquisa — aguarda e sinaliza para retentar
            self.logger.log(
                f"Auto {numero_auto}: DOM atualizado durante pesquisa (stale) — será retentado.", "WARNING"
            )
            time.sleep(0.8)
            return False
        except Exception as e:
            self.logger.log(f"Erro ao pesquisar auto {numero_auto}: {str(e)}", "ERROR")
            return False
    
    def verificar_resultado_pesquisa(self):
        """Confere se a pesquisa trouxe um resultado aproveitável."""
        from selenium.common.exceptions import StaleElementReferenceException

        def _buscar_checkboxes():
            return self.driver.find_elements(
                By.CSS_SELECTOR, "[id^='Corpo_gdvAutoInfracao_ckSelecionar_']"
            )

        try:
            # Espera a tela responder de um jeito ou de outro.
            try:
                def _grid_ou_nenhum(d):
                    try:
                        if d.find_elements(By.CSS_SELECTOR, "[id^='Corpo_gdvAutoInfracao_ckSelecionar_']"):
                            return True
                        return "Nenhum registro" in self._obter_texto_body(d)
                    except StaleElementReferenceException:
                        return False  # DOM em atualização, continua polling
                WebDriverWait(self.driver, 5, poll_frequency=0.3).until(_grid_ou_nenhum)
            except TimeoutException:
                pass

            # Em postback o DOM às vezes reaparece, então vale tentar algumas vezes.
            checkboxes = []
            for tentativa in range(3):
                try:
                    checkboxes = _buscar_checkboxes()
                    # Testa se os elementos estão "vivos" acessando uma propriedade
                    _ = [cb.is_enabled() for cb in checkboxes]
                    break
                except StaleElementReferenceException:
                    if tentativa < 2:
                        time.sleep(0.5)
                    else:
                        try:
                            checkboxes = _buscar_checkboxes()
                        except Exception:
                            checkboxes = []

            # Verificar se há erro de servidor
            if self.verificar_erro_servidor():
                self.logger.log("Erro de servidor detectado ao verificar resultado!", "ERROR")
                return "erro_servidor", 0

            # Sem checkbox, trata como não encontrado.
            if not checkboxes:
                texto_pagina = self._obter_texto_body()
                if "Nenhum registro" in texto_pagina:
                    return "nao_encontrado", 0
                return "nao_encontrado", 0

            quantidade = len(checkboxes)
            if quantidade == 1:
                return "encontrado", 1
            else:
                return "multiplos", quantidade

        except Exception as e:
            self.logger.log(f"Erro ao verificar resultado: {str(e)}", "ERROR")
            return "erro", 0
    
    def _obter_checkbox_primeira_linha_dados(self):
        """Pega o checkbox da primeira linha real de dados, nunca o do cabeçalho."""
        # Primeiro tenta pelo id exato, que é o caminho mais rápido.
        try:
            el = self.driver.find_element(By.ID, "Corpo_gdvAutoInfracao_ckSelecionar_0")
            if el.is_displayed():
                return el
        except NoSuchElementException:
            pass
        # Se não achar, cai para a busca focada na tabela.
        try:
            el = self.driver.find_element(
                By.XPATH,
                "//*[@id='Corpo_gdvAutoInfracao']//tbody//input[contains(@id,'ckSelecionar_')]"
            )
            if el.is_displayed():
                return el
        except NoSuchElementException:
            pass
        return None

    def _ler_identificador_primeira_linha(self) -> str:
        """Lê o identificador do débito (ex: FRMEV01960242025) da primeira linha de dados.
        Tenta múltiplas estratégias. Retorna string vazia apenas se realmente não encontrar."""
        try:
            resultado = self.driver.execute_script("""
                // Estratégia 1: id padrão GridView gerado pelo ASP.NET
                var ids = [
                    'Corpo_gdvAutoInfracao_lnkIdentificadorDebito_0',
                    'ContentPlaceHolderCorpo_gdvAutoInfracao_lnkIdentificadorDebito_0',
                    'Corpo_gdvAutoInfracao_Label1_0',
                    'Corpo_gdvAutoInfracao_lnkDebito_0'
                ];
                for (var k = 0; k < ids.length; k++) {
                    var el = document.getElementById(ids[k]);
                    if (el) { var t = el.textContent.trim(); if (t.length > 3) return t; }
                }

                // Estratégia 2: qualquer link ou célula dentro da primeira linha de dados da grid
                var grid = document.getElementById('Corpo_gdvAutoInfracao');
                if (grid) {
                    var rows = grid.querySelectorAll('tbody tr');
                    for (var r = 0; r < rows.length; r++) {
                        var row = rows[r];
                        // Pular linhas de cabeçalho ou paginação (sem <td>)
                        var cells = row.querySelectorAll('td');
                        if (cells.length < 2) continue;

                        // Procurar link com texto longo (identificador)
                        var links = row.querySelectorAll('a');
                        for (var i = 0; i < links.length; i++) {
                            var txt = links[i].textContent.trim();
                            if (txt.length > 5) return txt;
                        }
                        // Sem link: pegar texto da segunda célula (primeira costuma ser checkbox)
                        var c1 = cells[1].textContent.trim();
                        if (c1.length > 5) return c1;
                    }
                }

                // Estratégia 3: campo de pesquisa — o valor digitado é o auto que foi buscado
                var campo = document.getElementById('Corpo_txbAutoInfracao');
                if (campo && campo.value && campo.value.trim().length > 3) return campo.value.trim();

                return '';
            """)
            return str(resultado).strip() if resultado else ''
        except Exception:
            return ''

    def _ler_auto_campo_pesquisa(self) -> str:
        """Lê o valor atual do campo de pesquisa (campo onde o número do auto foi digitado)."""
        try:
            val = self.driver.execute_script(
                "var f=document.getElementById('Corpo_txbAutoInfracao'); return f ? f.value.trim() : '';"
            )
            return str(val).strip() if val else ''
        except Exception:
            return ''

    def _aguardar_checkbox_marcado(self, timeout=4):
        """Aguarda o checkbox estar marcado (polling via JS). Sai assim que estiver checked — mais rápido que sleep fixo."""
        try:
            def _esta_marcado(driver):
                try:
                    return driver.execute_script(
                        "var cb=document.querySelector('#Corpo_gdvAutoInfracao_ckSelecionar_0');"
                        "return !!(cb && cb.checked);"
                    )
                except Exception:
                    return False
            WebDriverWait(self.driver, timeout, poll_frequency=0.2).until(_esta_marcado)
            return True
        except TimeoutException:
            return False

    def _ler_estado_checkbox(self) -> bool:
        """Lê o estado atual do checkbox via JS. Retorna True se marcado, False se desmarcado ou ausente."""
        try:
            return bool(self.driver.execute_script(
                "var cb=document.querySelector('#Corpo_gdvAutoInfracao_ckSelecionar_0');"
                "return !!(cb && cb.checked);"
            ))
        except Exception:
            return False

    def _clicar_checkbox_auto(self):
        """Clica no checkbox via JS com verificação estrita de mudança de estado.

        Regra de ouro: só retorna True se o estado mudou de False → True.
        Se o checkbox já estiver marcado ao entrar neste método, retorna False — o
        chamador (processar_autos PASSO B) já tratou esse caso antes de chegar aqui,
        portanto um checkbox marcado neste ponto indica condição inesperada.
        """
        try:
            self._aguardar_overlay_sumir(timeout=5)
            time.sleep(0.25)  # Estabilizar antes de ler o estado

            # Passo 1 — confirmar que o checkbox está desmarcado antes de clicar
            if self._ler_estado_checkbox():
                # Marcado inesperadamente neste ponto — não clicar para evitar desmarcar
                return False

            # Passo 2 — executar o clique
            clicou = self.driver.execute_script(
                "var b=document.getElementById('wings_process_presentation_dashboard_bar');"
                "if(b)b.style.display='none';"
                "var cb=document.querySelector('#Corpo_gdvAutoInfracao_ckSelecionar_0');"
                "if(!cb)return false;"
                "cb.click();"
                "return true;"
            )
            if not clicou:
                return False

            # Passo 3 — aguardar mudança desmarcado → marcado (até 4 s)
            if not self._aguardar_checkbox_marcado(timeout=4):
                return False

            # Passo 4 — confirmação final do estado
            return self._ler_estado_checkbox()

        except Exception:
            return False

    def _checkbox_foi_validado(self, max_tentativas=4, intervalo=0.35):
        """Verifica com retentativas se o checkbox continua marcado após o clique.
        Só retorna True se realmente estiver checked em ao menos uma das tentativas."""
        for _ in range(max_tentativas):
            try:
                time.sleep(intervalo)
                if self._ler_estado_checkbox():
                    return True
            except Exception:
                pass
        return False

    def _checkbox_ainda_marcado_apos_delay(self, delay=0.4):
        """Confirmação final: checkbox deve continuar marcado após um delay adicional."""
        try:
            time.sleep(delay)
            return self._ler_estado_checkbox()
        except Exception:
            return False
    
    def selecionar_e_inscrever(self):
        """Seleciona o checkbox da primeira linha de dados e clica em Incluir na SERASA"""
        try:
            # Clicar no checkbox via JS puro (fallback: tentar novamente se retornar false)
            if not self._clicar_checkbox_auto():
                self.driver.execute_script(
                    "var b=document.getElementById('wings_process_presentation_dashboard_bar');"
                    "if(b)b.style.display='none';"
                    "var cb=document.querySelector('#Corpo_gdvAutoInfracao_ckSelecionar_0');"
                    "if(cb)cb.click();"
                )
            self.logger.log("Checkbox selecionado!", "INFO")
            
            # Clicar em Incluir na SERASA
            botao_incluir = self.wait.until(
                EC.element_to_be_clickable((By.XPATH, '//*[@id="Corpo_btnIncluirSerasa"]'))
            )
            botao_incluir.click()
            self.logger.log("Clicado em 'Incluir na SERASA'...", "INFO")
            
            return True
            
        except Exception as e:
            self.logger.log(f"Erro ao selecionar/inscrever: {str(e)}", "ERROR")
            return False
    
    def aguardar_inscricao_completa(self):
        """Aguarda a área sumir após incluir na SERASA"""
        try:
            self.logger.log("Aguardando conclusão da inscrição...", "INFO")
            
            # Aguardar até 30 segundos
            max_tentativas = 30
            for tentativa in range(max_tentativas):
                time.sleep(1)
                
                # Verificar se a área de resultados sumiu
                try:
                    resultado = self.driver.find_element(By.XPATH, '//*[@id="Corpo_gdvAutoInfracao_ckSelecionar_0"]')
                    # Se ainda existe, continuar aguardando
                except NoSuchElementException:
                    # Área sumiu, inscrição concluída
                    self.logger.log("Inscrição concluída!", "SUCCESS")
                    return True
                
                # Verificar se há popup ou mensagem de erro
                try:
                    texto_pagina = self._obter_texto_body()
                    if "erro" in texto_pagina.lower() or "error" in texto_pagina.lower():
                        self.logger.log("Possível erro detectado na página", "WARNING")
                except:
                    pass
            
            # Se não sumiu, tentar F5
            self.logger.log("Área não sumiu, recarregando página...", "WARNING")
            self.driver.refresh()
            time.sleep(3)
            
            # Verificar novamente
            try:
                self.driver.find_element(By.XPATH, '//*[@id="Corpo_gdvAutoInfracao_ckSelecionar_0"]')
                # Ainda existe, tentar novamente
                self.logger.log("Tentando incluir novamente...", "WARNING")
                return False
            except NoSuchElementException:
                return True
                
        except Exception as e:
            self.logger.log(f"Erro ao aguardar inscrição: {str(e)}", "ERROR")
            return False
    
    def limpar_formulario(self):
        """Limpa a tela para o próximo auto."""
        try:
            botao_limpar = self.wait.until(
                EC.element_to_be_clickable((By.XPATH, '//*[@id="Corpo_btnLimpar"]'))
            )
            botao_limpar.click()
            time.sleep(0.4)
            self.logger.log("Formulário limpo!", "INFO")
            return True
        except Exception as e:
            self.logger.log(f"Erro ao limpar formulário: {str(e)}", "ERROR")
            return False
    
    def _verificar_sessao_expirada(self) -> bool:
        """Detecta se a sessão voltou para a tela de login."""
        try:
            url = self.driver.current_url or ""
            return "Login" in url or "login" in url.lower()
        except Exception:
            return False

    def processar_autos(self, autos, progress_callback=None, stats_callback=None,
                        error_handler=None, checkpoint: "CheckpointManager | None" = None,
                        idx_inicio: int = 0):
        """Percorre os autos e deixa a seleção pronta, sem concluir a inclusão."""
        total = len(autos)
        sucessos = 0
        erros = 0

        # Se a execução estiver sendo retomada, reaproveita o que já tinha sido salvo.
        if checkpoint and checkpoint.existe() and idx_inicio > 0:
            dados_cp = checkpoint.carregar()
            resultados_anteriores = dados_cp.get("resultados", [])
            for r in resultados_anteriores:
                self.resultados.append(r)
                if r.get("situacao") == "SELECIONADO":
                    sucessos += 1
                else:
                    erros += 1

        for idx in range(idx_inicio, total):
            auto = str(autos[idx]).strip()
            if self.parar:
                break
            if not auto:
                continue

            # Se a sessão cair, tenta restaurar antes de perder o andamento.
            if self._verificar_sessao_expirada():
                self.logger.log("Sessão expirada detectada — refazendo login automaticamente...", "WARNING")
                if self.usuario_login and self.senha_login:
                    if self.fazer_login(self.usuario_login, self.senha_login):
                        if not self.navegar_para_formulario():
                            self.logger.log("Não foi possível renavegar após sessão expirada. Interrompendo.", "ERROR")
                            break
                        self.logger.log("Sessão restaurada com sucesso.", "SUCCESS")
                    else:
                        self.logger.log("Falha ao refazer login após sessão expirada. Interrompendo.", "ERROR")
                        break
                else:
                    self.logger.log("Credenciais não disponíveis para restaurar sessão. Interrompendo.", "ERROR")
                    break
            
            try:
                if progress_callback:
                    progress_callback(f"Processando {idx+1}/{total}: {auto}")

                # Um intervalo curto entre autos ajuda o portal a responder melhor.
                if idx > idx_inicio:
                    time.sleep(random.uniform(0.8, 1.5))
                
                # Se a primeira pesquisa falhar, ainda tenta uma segunda vez.
                if not self.pesquisar_auto(auto):
                    self.logger.log(f"Auto {auto}: primeira tentativa de pesquisa falhou — retentando uma vez...", "WARNING")
                    time.sleep(0.6)
                    if not self.pesquisar_auto(auto):
                        self.logger.log(f"Auto {auto}: Erro ao pesquisar (timeout/conexão/stale)", "ERROR")
                        self.resultados.append({'auto': auto, 'situacao': 'ERRO AO PESQUISAR', 'horario': _ts()})
                        erros += 1
                        if stats_callback:
                            stats_callback(sucessos, erros)
                        try:
                            self.navegar_para_formulario()
                        except Exception:
                            pass
                        if idx + 1 < total:
                            self._preparar_proximo_auto(autos[idx + 1])
                        continue
                    self.logger.log(f"Auto {auto}: pesquisa OK na retentativa.", "INFO")
                
                # Com a pesquisa concluída, interpreta o retorno da tela.
                status, quantidade = self.verificar_resultado_pesquisa()
                
                if status == "erro_servidor":
                    self.logger.log(f"Auto {auto}: Erro de servidor detectado", "ERROR")
                    # Quando o portal quebra, tenta se recuperar e seguir do mesmo ponto.
                    if self.tratar_erro_servidor(tentar_navegar_novamente=True):
                        if not self.navegar_para_formulario():
                            self.logger.log("Não foi possível navegar após tratar erro. Interrompendo...", "ERROR")
                            break
                        # Retentar pesquisa do mesmo auto
                        self.logger.log(f"Retentando pesquisa do auto {auto} após tratar erro...", "INFO")
                        resultado_pesquisa = self.pesquisar_auto(auto)
                        if resultado_pesquisa:
                            # Verificar resultado novamente
                            status, quantidade = self.verificar_resultado_pesquisa()
                            if status == "erro_servidor":
                                # Se ainda tiver erro, marcar e continuar
                                self.logger.log(f"Auto {auto}: Erro de servidor persiste após retentar!", "ERROR")
                                self.resultados.append({'auto': auto, 'situacao': 'ERRO DE SERVIDOR', 'horario': _ts()})
                                erros += 1
                                if stats_callback:
                                    stats_callback(sucessos, erros)
                                if idx + 1 < total:
                                    self._preparar_proximo_auto(autos[idx + 1])
                                continue
                            # Se não tiver mais erro, processar normalmente (cair no código abaixo)
                            # Não fazer continue aqui, deixar o código processar normalmente
                        else:
                            # Se falhou ao pesquisar, marcar como erro
                            self.logger.log(f"Auto {auto}: Falha ao pesquisar após tratar erro!", "ERROR")
                            self.resultados.append({'auto': auto, 'situacao': 'ERRO DE SERVIDOR', 'horario': _ts()})
                            erros += 1
                            if stats_callback:
                                stats_callback(sucessos, erros)
                            if idx + 1 < total:
                                self._preparar_proximo_auto(autos[idx + 1])
                            continue
                    else:
                        # Se falhou ao tratar erro, marcar e continuar
                        self.logger.log(f"Auto {auto}: Falha ao tratar erro de servidor!", "ERROR")
                        self.resultados.append({'auto': auto, 'situacao': 'ERRO DE SERVIDOR', 'horario': _ts()})
                        erros += 1
                        if stats_callback:
                            stats_callback(sucessos, erros)
                        if idx + 1 < total:
                            self._preparar_proximo_auto(autos[idx + 1])
                        continue
                
                # Se chegou aqui e status ainda é erro_servidor, pular processamento
                if status == "erro_servidor":
                    continue
                
                if status == "nao_encontrado":
                    self.logger.log(f"Auto {auto}: Não encontrado na caixa", "WARNING")
                    self.resultados.append({'auto': auto, 'situacao': 'NÃO ENCONTRADO NA CAIXA', 'horario': _ts()})
                    erros += 1
                    if stats_callback:
                        stats_callback(sucessos, erros)
                    
                    # Passar para próximo auto usando CTRL+A e digitar próximo
                    if idx + 1 < total:
                        self._preparar_proximo_auto(autos[idx + 1])
                    continue
                
                elif status == "multiplos":
                    self.logger.log(f"Auto {auto}: Múltiplos resultados encontrados ({quantidade})", "WARNING")
                    self.resultados.append({'auto': auto, 'situacao': 'MÚLTIPLOS RESULTADOS', 'horario': _ts()})
                    erros += 1
                    if stats_callback:
                        stats_callback(sucessos, erros)
                    
                    # Passar para próximo auto usando CTRL+A e digitar próximo
                    if idx + 1 < total:
                        self._preparar_proximo_auto(autos[idx + 1])
                    continue
                
                elif status == "encontrado":
                    # ── Seleção do checkbox com verificação de identidade + mudança de estado ──
                    try:
                        # ── PASSO A: Confirmar que a grid mostra o auto correto ──
                        # Quando outra janela fica na frente, o Chrome pode atrasar a atualização
                        # do DOM mesmo após o postback concluir. Por isso, se o identificador
                        # não bater, re-pesquisamos o auto em vez de simplesmente aguardar.
                        ja_selecionado = False
                        identificador_ok = False
                        auto_upper = auto.strip().upper()

                        def _confirmar_identidade() -> bool:
                            """Lê o identificador da grid (ou campo de busca) e verifica se bate
                            com o auto pesquisado. Retorna True se confirmado ou se não foi possível
                            ler nenhum identificador (benefício da dúvida)."""
                            nonlocal ja_selecionado
                            ident = self._ler_identificador_primeira_linha()
                            if not ident:
                                ident = self._ler_auto_campo_pesquisa()
                            estado = self._ler_estado_checkbox()
                            if not ident:
                                # Sem identificador legível — benefício da dúvida
                                ja_selecionado = estado
                                return True
                            if auto_upper in ident.strip().upper():
                                ja_selecionado = estado
                                return True
                            return False

                        # Tentativa 1: identificador logo após a pesquisa (já aguardou 2 s)
                        if _confirmar_identidade():
                            identificador_ok = True
                        else:
                            # Grid ainda mostra auto anterior — aguardar mais 1.5 s e tentar de novo
                            self.logger.log(
                                f"Auto {auto}: Grid ainda mostra resultado anterior — aguardando mais.", "WARNING"
                            )
                            time.sleep(1.5)
                            if _confirmar_identidade():
                                identificador_ok = True
                            else:
                                # DOM persiste desatualizado — RE-PESQUISAR o auto
                                self.logger.log(
                                    f"Auto {auto}: Re-pesquisando para forçar atualização da grid.", "WARNING"
                                )
                                if self.pesquisar_auto(auto):
                                    time.sleep(1.0)
                                    if _confirmar_identidade():
                                        identificador_ok = True
                                    else:
                                        self.logger.log(
                                            f"Auto {auto}: Grid ainda divergente após re-pesquisa. Pulando.", "ERROR"
                                        )
                                else:
                                    self.logger.log(
                                        f"Auto {auto}: Re-pesquisa falhou. Pulando.", "ERROR"
                                    )

                        if not identificador_ok:
                            self.resultados.append({'auto': auto, 'situacao': 'RESULTADO INCORRETO NA GRID', 'horario': _ts()})
                            erros += 1
                            if stats_callback:
                                stats_callback(sucessos, erros)
                            if idx + 1 < total:
                                self._preparar_proximo_auto(autos[idx + 1])
                            continue

                        # ── PASSO B: Auto já estava selecionado — contar como sucesso ──
                        if ja_selecionado:
                            self.logger.log(
                                f"Auto {auto}: Já estava selecionado no sistema. Contabilizado como sucesso.", "SUCCESS"
                            )
                            self.resultados.append({'auto': auto, 'situacao': 'SELECIONADO', 'horario': _ts()})
                            sucessos += 1
                            if stats_callback:
                                stats_callback(sucessos, erros)
                            if idx + 1 < total:
                                self._preparar_proximo_auto(autos[idx + 1])
                            continue

                        # ── PASSO C: Clicar no checkbox (checkbox está desmarcado e grid é o auto correto) ──
                        clique_ok = self._clicar_checkbox_auto()

                        if not clique_ok:
                            # Fallback: só clicar se ainda estiver desmarcado
                            if not self._ler_estado_checkbox():
                                self.logger.log(f"Auto {auto}: Clique principal falhou — tentando fallback.", "WARNING")
                                self.driver.execute_script(
                                    "var b=document.getElementById('wings_process_presentation_dashboard_bar');"
                                    "if(b)b.style.display='none';"
                                    "var cb=document.querySelector('#Corpo_gdvAutoInfracao_ckSelecionar_0');"
                                    "if(cb && !cb.checked)cb.click();"
                                )
                                time.sleep(0.6)

                        # ── PASSO D: Validação dupla do estado final ──
                        if not self._checkbox_foi_validado():
                            self.logger.log(f"Auto {auto}: Clique no checkbox não foi validado pelo sistema (não contou).", "WARNING")
                            self.resultados.append({'auto': auto, 'situacao': 'CLIQUE NÃO VALIDADO', 'horario': _ts()})
                            erros += 1
                            if stats_callback:
                                stats_callback(sucessos, erros)
                        elif not self._checkbox_ainda_marcado_apos_delay(delay=0.3):
                            self.logger.log(f"Auto {auto}: Checkbox não permaneceu marcado após confirmação (não contou).", "WARNING")
                            self.resultados.append({'auto': auto, 'situacao': 'CLIQUE NÃO CONFIRMADO', 'horario': _ts()})
                            erros += 1
                            if stats_callback:
                                stats_callback(sucessos, erros)
                        else:
                            self.logger.log(f"Auto {auto}: Checkbox selecionado e validado!", "SUCCESS")
                            self.resultados.append({'auto': auto, 'situacao': 'SELECIONADO', 'horario': _ts()})
                            sucessos += 1
                            if stats_callback:
                                stats_callback(sucessos, erros)
                    except Exception as e:
                        self.logger.log(f"Auto {auto}: Erro ao selecionar checkbox: {str(e)}", "ERROR")
                        self.resultados.append({'auto': auto, 'situacao': 'ERRO AO SELECIONAR', 'horario': _ts()})
                        erros += 1
                        if stats_callback:
                            stats_callback(sucessos, erros)
                    
                    # Passar para próximo auto usando CTRL+A e digitar próximo
                    if idx + 1 < total:
                        self._preparar_proximo_auto(autos[idx + 1])
                
            except Exception as e:
                self.logger.log(f"Erro ao processar auto {auto}: {str(e)}", "ERROR")
                self.resultados.append({'auto': auto, 'situacao': 'ERRO', 'horario': _ts()})
                erros += 1
                if stats_callback:
                    stats_callback(sucessos, erros)
                
                # Verificar se precisa pausar
                if error_handler:
                    acao = error_handler(auto, str(e))
                    if acao == "pausar":
                        self.pausado = True
                        self.logger.log("Automação pausada pelo usuário", "WARNING")
                        self.aguardar_pausa()
                    elif acao == "parar":
                        self.parar = True
                        break
                
                # Passar para próximo auto usando CTRL+A e digitar próximo
                if idx + 1 < total:
                    try:
                        self._preparar_proximo_auto(autos[idx + 1])
                    except:
                        pass

            # ── Melhoria 1: salvar checkpoint após cada auto ──
            if checkpoint:
                checkpoint.salvar(idx + 1, total, list(self.resultados))
        
        # Checkpoint concluído — limpar arquivo
        if checkpoint:
            checkpoint.limpar()

        return sucessos, erros
    
    def _preparar_proximo_auto(self, proximo_auto):
        """Prepara o campo para o próximo auto via JS — sem sleeps intermediários."""
        try:
            self.driver.execute_script(
                "var f=document.getElementById('Corpo_txbAutoInfracao');"
                "if(f){f.value=arguments[0];f.dispatchEvent(new Event('input',{bubbles:true}));}",
                str(proximo_auto)
            )
            self.logger.log(f"Campo preparado para próximo auto: {proximo_auto}", "INFO")
        except Exception as e:
            self.logger.log(f"Erro ao preparar próximo auto: {str(e)}", "WARNING")
    
    def salvar_resultados(self, caminho_original, sufixo_arquivo=None):
        """Salva os resultados na planilha. sufixo_arquivo: ex. 'Reprocessamento' para nome diferente."""
        try:
            from datetime import datetime as _dt
            caminho = Path(caminho_original)
            nome_extra = f" - {sufixo_arquivo}" if sufixo_arquivo else ""
            data_hora = _dt.now().strftime("%d-%m-%Y %Hh%M")
            dados = []
            for resultado in self.resultados:
                dados.append({
                    'Autos de infração': resultado['auto'],
                    'Situação': resultado['situacao'],
                    'Horário': resultado.get('horario', '')
                })
            df_resultado = pd.DataFrame(dados)
            if caminho.suffix.lower() == '.csv':
                novo_caminho = caminho.parent / f"Planilha Inscrição Serasa Resultado{nome_extra} {data_hora}.csv"
                df_resultado.to_csv(novo_caminho, index=False, encoding='utf-8')
            else:
                novo_caminho = caminho.parent / f"Planilha Inscrição Serasa Resultado{nome_extra} {data_hora}.xlsx"
                df_resultado.to_excel(novo_caminho, index=False, engine='openpyxl')
                self._aplicar_formatacao_excel(novo_caminho, 'Situação')
            self.logger.log(f"Resultados salvos em: {novo_caminho}", "SUCCESS")
            return novo_caminho
        except Exception as e:
            self.logger.log(f"Erro ao salvar resultados: {str(e)}", "ERROR")
            return None

    def salvar_log_excel(self, caminho_planilha: str):
        """Exporta todo o histórico de logs para uma aba extra na planilha de resultados."""
        try:
            if not caminho_planilha or not caminho_planilha.endswith(".xlsx"):
                return
            from datetime import datetime as _dt
            wb = load_workbook(caminho_planilha)
            nome_aba = "Log"
            if nome_aba in wb.sheetnames:
                del wb[nome_aba]
            ws_log = wb.create_sheet(nome_aba)
            # Cabeçalho
            fill_hdr = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            font_hdr = Font(bold=True, color="FFFFFF", size=10)
            for col_idx, titulo in enumerate(["Horário", "Tipo", "Mensagem"], start=1):
                c = ws_log.cell(row=1, column=col_idx, value=titulo)
                c.fill = fill_hdr
                c.font = font_hdr
                c.alignment = Alignment(horizontal="center", vertical="center")
            ws_log.column_dimensions["A"].width = 22
            ws_log.column_dimensions["B"].width = 12
            ws_log.column_dimensions["C"].width = 80
            # Mapeamento de cores por tipo
            cores_tipo = {
                "SUCCESS": "C6EFCE",
                "ERROR":   "FFC7CE",
                "WARNING": "FFEB9C",
                "INFO":    "DDEBF7",
            }
            for row_idx, linha in enumerate(self.logger.logs, start=2):
                # Formato esperado: [HH:MM:SS] [TIPO] Mensagem
                partes = linha.strip().split("] ", 2)
                horario = partes[0].lstrip("[") if len(partes) >= 1 else ""
                tipo = partes[1].lstrip("[") if len(partes) >= 2 else ""
                mensagem = partes[2] if len(partes) >= 3 else linha
                ws_log.cell(row=row_idx, column=1, value=horario)
                ws_log.cell(row=row_idx, column=2, value=tipo)
                c_msg = ws_log.cell(row=row_idx, column=3, value=mensagem)
                cor = cores_tipo.get(tipo, "FFFFFF")
                fill_row = PatternFill(start_color=cor, end_color=cor, fill_type="solid")
                for col in range(1, 4):
                    ws_log.cell(row=row_idx, column=col).fill = fill_row
            wb.save(caminho_planilha)
            self.logger.log("Log exportado para aba 'Log' na planilha de resultados.", "INFO")
        except Exception as e:
            self.logger.log(f"Erro ao salvar log em Excel: {str(e)}", "WARNING")

    def _aplicar_formatacao_excel(self, caminho_excel, nome_coluna_situacao):
        """Aplica formatação de tabela, cabeçalho colorido e cores por status na planilha Excel"""
        # Mapeamento de cores por situação
        _CORES_STATUS = {
            "SELECIONADO":           ("C6EFCE", "375623"),  # verde
            "ERRO":                  ("FFC7CE", "9C0006"),  # vermelho
            "CLIQUE NÃO VALIDADO":   ("FFEB9C", "9C5700"),  # amarelo
            "CLIQUE NÃO CONFIRMADO": ("FFEB9C", "9C5700"),  # amarelo
            "ERRO AO SELECIONAR":    ("FFC7CE", "9C0006"),  # vermelho
            "AUTO NÃO ENCONTRADO":   ("FCE4D6", "843C0C"),  # laranja
        }
        try:
            wb = load_workbook(caminho_excel)
            ws = wb.active
            
            # Definir largura das colunas (A=auto, B=situação, C=horário)
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 28
            ws.column_dimensions['C'].width = 20
            
            # Formatar cabeçalho (linha 1)
            fill_cabecalho = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            font_cabecalho = Font(bold=True, color="FFFFFF", size=11)
            alignment_cabecalho = Alignment(horizontal="center", vertical="center")
            
            n_cols = ws.max_column
            for col in range(1, n_cols + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = fill_cabecalho
                cell.font = font_cabecalho
                cell.alignment = alignment_cabecalho
            
            # Adicionar bordas
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Localizar coluna de situação
            col_situacao = None
            for col in range(1, n_cols + 1):
                if ws.cell(row=1, column=col).value == nome_coluna_situacao:
                    col_situacao = col
                    break

            # Aplicar bordas e cores por status
            max_row = ws.max_row
            for row in range(1, max_row + 1):
                for col in range(1, n_cols + 1):
                    ws.cell(row=row, column=col).border = thin_border
                # Colorir linha de dados com base no status
                if row > 1 and col_situacao:
                    status = str(ws.cell(row=row, column=col_situacao).value or "")
                    if status in _CORES_STATUS:
                        bg, fg = _CORES_STATUS[status]
                        fill_row = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
                        font_row = Font(color=fg, size=10)
                        for col in range(1, n_cols + 1):
                            ws.cell(row=row, column=col).fill = fill_row
                            ws.cell(row=row, column=col).font = font_row
            
            # Criar tabela (formato de tabela Excel)
            if max_row > 1:
                col_letra = "ABC"[n_cols - 1] if n_cols <= 3 else "C"
                tab = Table(displayName="TabelaResultados", ref=f"A1:{col_letra}{max_row}")
                style = TableStyleInfo(
                    name="TableStyleMedium9",
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False
                )
                tab.tableStyleInfo = style
                ws.add_table(tab)
            
            # Centralizar dados nas células
            for row in range(2, max_row + 1):
                for col in range(1, n_cols + 1):
                    ws.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center")
            
            wb.save(caminho_excel)
            
        except Exception as e:
            self.logger.log(f"Erro ao aplicar formatação: {str(e)}", "WARNING")


class InterfaceGrafica:
    """Interface gráfica moderna e integrada"""
    
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Automação SIFAMA - ANTT")
        self.root.geometry("450x350")  # Menor para tela de login
        self.root.resizable(False, False)
        
        # Variáveis
        self.tipo_automacao = tk.StringVar(value="consulta")
        self.planilha_path = None
        self.autos = []
        self.automacao = None
        self.automacao_temp = None  # Automação temporária do login (mantém navegador aberto)
        self.thread_automacao = None
        self.sucessos = 0
        self.erros = 0
        self.erro_dialogo_ativo = False
        self.logado = False
        self.usuario_logado = None
        self.senha_logada = None
        self.ultimos_autos_com_erro = []  # Para botão "Reprocessar apenas erros"
        
        # Logger
        self.logger = Logger(self.atualizar_log)
        
        # Frames principais
        self.frame_login = None
        self.frame_principal = None
        
        self.criar_interface()
    
    def criar_interface(self):
        """Cria a interface gráfica com telas separadas"""
        # Estilo
        estilo = ttk.Style()
        estilo.theme_use('clam')
        
        # Container principal
        self.main_container = tk.Frame(self.root, bg='#f0f0f0')
        self.main_container.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Criar tela de login
        self.criar_tela_login()
        
        # Criar tela principal (inicialmente escondida)
        self.criar_tela_principal()
        
        # Mostrar tela de login inicialmente
        self.mostrar_tela_login()
    
    def criar_tela_login(self):
        """Cria a tela de login - simples e compacta"""
        self.frame_login = tk.Frame(self.main_container, bg='#f0f0f0')
        
        # Título simples
        titulo = tk.Label(self.frame_login, text="Automação SIFAMA - ANTT", 
                         font=("Arial", 16, "bold"), bg='#f0f0f0', fg='#2c3e50')
        titulo.pack(pady=30)
        
        # Frame de login compacto
        frame_login_box = tk.Frame(self.frame_login, bg='#ffffff', relief=tk.RAISED, bd=1)
        frame_login_box.pack(padx=100, pady=20)
        
        login_inner = tk.Frame(frame_login_box, bg='#ffffff')
        login_inner.pack(padx=30, pady=25)
        
        # Usuário
        tk.Label(login_inner, text="Usuário:", font=("Arial", 10), 
                bg='#ffffff').grid(row=0, column=0, sticky="w", pady=8, padx=5)
        self.entry_usuario = tk.Entry(login_inner, width=25, font=("Arial", 10))
        self.entry_usuario.grid(row=0, column=1, pady=8, padx=5, ipady=3)
        self.entry_usuario.focus()
        
        # Senha
        tk.Label(login_inner, text="Senha:", font=("Arial", 10), 
                bg='#ffffff').grid(row=1, column=0, sticky="w", pady=8, padx=5)
        self.entry_senha = tk.Entry(login_inner, width=25, show="*", font=("Arial", 10))
        self.entry_senha.grid(row=1, column=1, pady=8, padx=5, ipady=3)
        
        # Botão Entrar
        btn_entrar_frame = tk.Frame(login_inner, bg='#ffffff')
        btn_entrar_frame.grid(row=2, column=0, columnspan=2, pady=15)
        
        self.btn_entrar = tk.Button(btn_entrar_frame, text="Entrar", 
                                    command=self.fazer_login, bg='#27ae60', fg='white',
                                    font=("Arial", 10, "bold"), width=15, height=1)
        self.btn_entrar.pack()
        
        # Bind Enter key
        self.entry_senha.bind('<Return>', lambda e: self.fazer_login())
        self.entry_usuario.bind('<Return>', lambda e: self.entry_senha.focus())
        
        # Status do login
        self.label_status_login = tk.Label(login_inner, text="", 
                                           font=("Arial", 9), bg='#ffffff', fg='#e74c3c')
        self.label_status_login.grid(row=3, column=0, columnspan=2, pady=5)
    
    def criar_tela_principal(self):
        """Cria a tela principal após login"""
        self.frame_principal = tk.Frame(self.main_container, bg='#f0f0f0')
        
        # Título
        titulo_frame = tk.Frame(self.frame_principal, bg='#2c3e50', height=60)
        titulo_frame.pack(fill=tk.X, pady=(0, 10))
        titulo_frame.pack_propagate(False)
        
        titulo_container = tk.Frame(titulo_frame, bg='#2c3e50')
        titulo_container.pack(fill=tk.BOTH, expand=True)
        
        titulo = tk.Label(titulo_container, text="Automação SIFAMA - ANTT", 
                         font=("Arial", 18, "bold"), bg='#2c3e50', fg='white')
        titulo.pack(side=tk.LEFT, padx=20, pady=15)
        
        # Label de usuário logado
        self.label_usuario_logado = tk.Label(titulo_container, text="", 
                                            font=("Arial", 10), bg='#2c3e50', fg='#ecf0f1')
        self.label_usuario_logado.pack(side=tk.RIGHT, padx=20)
        
        # Botão sair
        btn_sair = tk.Button(titulo_container, text="Sair", 
                            command=self.sair, bg='#e74c3c', fg='white',
                            font=("Arial", 9), relief=tk.RAISED, cursor='hand2', width=10)
        btn_sair.pack(side=tk.RIGHT, padx=10)
        
        # Frame de seleção de automação
        frame_tipo = tk.LabelFrame(self.frame_principal, text="Tipo de Automação", 
                                   font=("Arial", 10, "bold"), bg='#f0f0f0')
        frame_tipo.pack(fill=tk.X, pady=5, padx=10)
        
        rb_consulta = tk.Radiobutton(frame_tipo, text="Consulta de Pagamento (Situação da Dívida)", 
                                     variable=self.tipo_automacao, value="consulta",
                                     font=("Arial", 9), bg='#f0f0f0', activebackground='#f0f0f0')
        rb_consulta.pack(side=tk.LEFT, padx=20, pady=10)
        
        rb_inscricao = tk.Radiobutton(frame_tipo, text="Inscrição na SERASA", 
                                     variable=self.tipo_automacao, value="inscricao",
                                     font=("Arial", 9), bg='#f0f0f0', activebackground='#f0f0f0')
        rb_inscricao.pack(side=tk.LEFT, padx=20, pady=10)
        
        # Frame de planilha
        frame_planilha = tk.LabelFrame(self.frame_principal, text="Planilha", 
                                       font=("Arial", 10, "bold"), bg='#f0f0f0')
        frame_planilha.pack(fill=tk.X, pady=5, padx=10)
        
        planilha_inner = tk.Frame(frame_planilha, bg='#f0f0f0')
        planilha_inner.pack(padx=10, pady=10)
        
        self.label_planilha = tk.Label(planilha_inner, text="Nenhuma planilha selecionada", 
                                       fg="gray", font=("Arial", 9), bg='#f0f0f0')
        self.label_planilha.pack(pady=5)
        
        btn_selecionar = tk.Button(planilha_inner, text="📁 Selecionar Planilha", 
                                   command=self.selecionar_planilha, bg='#3498db', fg='white',
                                   font=("Arial", 9), relief=tk.RAISED, cursor='hand2', width=20)
        btn_selecionar.pack(pady=5)
        
        # Frame de estatísticas
        frame_stats = tk.Frame(self.frame_principal, bg='#ecf0f1', relief=tk.RAISED, bd=2)
        frame_stats.pack(fill=tk.X, pady=5, padx=10)
        
        stats_inner = tk.Frame(frame_stats, bg='#ecf0f1')
        stats_inner.pack(padx=10, pady=10)
        
        tk.Label(stats_inner, text="Sucessos:", font=("Arial", 10, "bold"), 
                bg='#ecf0f1', fg='#27ae60').pack(side=tk.LEFT, padx=10)
        self.label_sucessos = tk.Label(stats_inner, text="0", font=("Arial", 12, "bold"), 
                                       bg='#ecf0f1', fg='#27ae60')
        self.label_sucessos.pack(side=tk.LEFT, padx=5)
        
        tk.Label(stats_inner, text="Erros:", font=("Arial", 10, "bold"), 
                bg='#ecf0f1', fg='#e74c3c').pack(side=tk.LEFT, padx=10)
        self.label_erros = tk.Label(stats_inner, text="0", font=("Arial", 12, "bold"), 
                                    bg='#ecf0f1', fg='#e74c3c')
        self.label_erros.pack(side=tk.LEFT, padx=5)
        
        # Frame de logs
        frame_logs = tk.LabelFrame(self.frame_principal, text="Logs em Tempo Real", 
                                   font=("Arial", 10, "bold"), bg='#f0f0f0')
        frame_logs.pack(fill=tk.BOTH, expand=True, pady=5, padx=10)
        
        # Text widget com scroll
        self.text_logs = scrolledtext.ScrolledText(frame_logs, height=15, width=80, 
                                                   font=("Consolas", 8), bg='#2c3e50', fg='#ecf0f1',
                                                   insertbackground='white')
        self.text_logs.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Frame de controles
        frame_controles = tk.Frame(self.frame_principal, bg='#f0f0f0')
        frame_controles.pack(fill=tk.X, pady=5, padx=10)
        
        self.btn_iniciar = tk.Button(frame_controles, text="▶ Iniciar Automação", 
                                     command=self.iniciar_automacao, bg='#27ae60', fg='white',
                                     font=("Arial", 10, "bold"), relief=tk.RAISED, 
                                     cursor='hand2', width=20, height=2)
        self.btn_iniciar.pack(side=tk.LEFT, padx=5)
        
        self.btn_pausar = tk.Button(frame_controles, text="⏸ Pausar", 
                                    command=self.pausar_automacao, bg='#f39c12', fg='white',
                                    font=("Arial", 10, "bold"), relief=tk.RAISED, 
                                    cursor='hand2', width=15, height=2, state=tk.DISABLED)
        self.btn_pausar.pack(side=tk.LEFT, padx=5)
        
        self.btn_continuar = tk.Button(frame_controles, text="▶ Continuar", 
                                      command=self.continuar_automacao, bg='#3498db', fg='white',
                                      font=("Arial", 10, "bold"), relief=tk.RAISED, 
                                      cursor='hand2', width=15, height=2, state=tk.DISABLED)
        self.btn_continuar.pack(side=tk.LEFT, padx=5)
        
        self.btn_parar = tk.Button(frame_controles, text="⏹ Parar", 
                                  command=self.parar_automacao, bg='#e74c3c', fg='white',
                                  font=("Arial", 10, "bold"), relief=tk.RAISED, 
                                  cursor='hand2', width=15, height=2, state=tk.DISABLED)
        self.btn_parar.pack(side=tk.LEFT, padx=5)
        
        self.btn_reprocessar_erros = tk.Button(frame_controles, text="🔄 Reprocessar apenas erros", 
                                               command=self.reprocessar_apenas_erros, bg='#9b59b6', fg='white',
                                               font=("Arial", 9, "bold"), relief=tk.RAISED, 
                                               cursor='hand2', state=tk.DISABLED)
        self.btn_reprocessar_erros.pack(side=tk.LEFT, padx=5)

        self.btn_gerar_planilha = tk.Button(frame_controles, text="💾 Gerar Planilha de Resultado",
                                            command=self.gerar_planilha_resultado, bg='#16a085', fg='white',
                                            font=("Arial", 9, "bold"), relief=tk.RAISED,
                                            cursor='hand2', state=tk.DISABLED)
        self.btn_gerar_planilha.pack(side=tk.LEFT, padx=5)

        self.btn_exportar_log = tk.Button(frame_controles, text="📋 Exportar Log",
                                          command=self.exportar_log_agora, bg='#7f8c8d', fg='white',
                                          font=("Arial", 9, "bold"), relief=tk.RAISED,
                                          cursor='hand2')
        self.btn_exportar_log.pack(side=tk.LEFT, padx=5)
        
        # Progresso
        self.progress_var = tk.StringVar(value="Aguardando...")
        self.label_progress = tk.Label(self.frame_principal, textvariable=self.progress_var, 
                                      fg="blue", font=("Arial", 9), bg='#f0f0f0')
        self.label_progress.pack(pady=5)
        
        self.progress_bar = ttk.Progressbar(self.frame_principal, mode='determinate', maximum=100)
        self.progress_bar.pack(fill=tk.X, pady=5, padx=10)
    
    def mostrar_tela_login(self):
        """Mostra a tela de login"""
        if self.frame_principal:
            self.frame_principal.pack_forget()
        if self.frame_login:
            self.frame_login.pack(fill=tk.BOTH, expand=True)
    
    def mostrar_tela_principal(self):
        """Mostra a tela principal"""
        # Ajustar tamanho da janela para tela principal
        self.root.geometry("900x700")
        self.root.resizable(True, True)
        
        if self.frame_login:
            self.frame_login.pack_forget()
        if self.frame_principal:
            self.frame_principal.pack(fill=tk.BOTH, expand=True)
    
    def fazer_login(self):
        """Realiza o login"""
        usuario = self.entry_usuario.get().strip()
        senha = self.entry_senha.get().strip()
        
        if not usuario or not senha:
            self.label_status_login.config(text="Preencha usuário e senha.", fg='#e74c3c')
            return
        
        # Desabilitar botão durante login
        self.btn_entrar.config(state=tk.DISABLED, text="Entrando...")
        self.label_status_login.config(text="Conectando...", fg='#3498db')
        self.root.update()
        
        # Criar automação temporária para testar login
        temp_automacao = BaseAutomacao(self.logger)
        
        def testar_login():
            try:
                print("DEBUG: Iniciando teste de login...")
                self.root.after(0, lambda: self.logger.log("Validando credenciais...", "INFO"))
                
                # Criar driver em modo headless (sem abrir janela) para validação
                print("DEBUG: Criando driver headless...")
                if not temp_automacao.criar_driver(headless=True):
                    print("DEBUG: Falha ao criar driver")
                    self.root.after(0, lambda: self.label_status_login.config(
                        text="Erro ao iniciar navegador. Verifique o ChromeDriver.", fg='#e74c3c'))
                    self.root.after(0, lambda: self.btn_entrar.config(state=tk.NORMAL, text="Entrar"))
                    return
                
                print("DEBUG: Driver criado, fazendo login...")
                resultado = temp_automacao.fazer_login(usuario, senha)
                print(f"DEBUG: Resultado do login: {resultado}")
                
                if resultado:
                    # Login bem-sucedido - NÃO FECHAR navegador (manter aberto)
                    print("DEBUG: Login bem-sucedido!")
                    self.logado = True
                    self.usuario_logado = usuario
                    self.senha_logada = senha
                    # Guardar referência do navegador headless (não fechar)
                    self.automacao_temp = temp_automacao
                    
                    self.root.after(0, lambda: self.logger.log("Login validado com sucesso!", "SUCCESS"))
                    self.root.after(0, lambda: self.label_usuario_logado.config(
                        text=f"Usuário: {usuario}"))
                    self.root.after(0, self.mostrar_tela_principal)
                    self.root.after(0, lambda: self.logger.log("Login realizado com sucesso!", "SUCCESS"))
                else:
                    # Login falhou - fechar navegador
                    print("DEBUG: Login falhou")
                    temp_automacao.fechar()
                    self.root.after(0, lambda: self.label_status_login.config(
                        text="Usuário ou senha inválidos. Tente novamente.", fg='#e74c3c'))
                    self.root.after(0, lambda: self.btn_entrar.config(state=tk.NORMAL, text="Entrar"))
                    
            except Exception as e:
                import traceback
                erro_completo = traceback.format_exc()
                print(f"ERRO NO LOGIN: {erro_completo}")  # Debug
                if temp_automacao.driver:
                    temp_automacao.fechar()
                self.root.after(0, lambda: self.label_status_login.config(
                    text=f"Erro: {str(e)}", fg='#e74c3c'))
                self.root.after(0, lambda: self.btn_entrar.config(state=tk.NORMAL, text="Entrar"))
                self.root.after(0, lambda: self.logger.log(f"Erro no login: {str(e)}", "ERROR"))
        
        # Executar login em thread
        thread_login = threading.Thread(target=testar_login, daemon=True)
        thread_login.start()
    
    def sair(self):
        """Sai do sistema e volta para tela de login"""
        # Fechar todas as automações
        if self.automacao and self.automacao.driver:
            self.automacao.fechar()
        if self.automacao_temp and self.automacao_temp.driver:
            self.automacao_temp.fechar()
        
        self.logado = False
        self.usuario_logado = None
        self.senha_logada = None
        self.planilha_path = None
        self.autos = []
        self.automacao = None
        self.automacao_temp = None
        self.label_planilha.config(text="Nenhuma planilha selecionada", fg="gray")
        self.text_logs.delete(1.0, tk.END)
        self.sucessos = 0
        self.erros = 0
        self._atualizar_stats_ui()
        
        # Ajustar tamanho da janela para tela de login
        self.root.geometry("450x350")
        self.root.resizable(False, False)
        
        self.mostrar_tela_login()
        self.entry_usuario.delete(0, tk.END)
        self.entry_senha.delete(0, tk.END)
        self.label_status_login.config(text="")
    
    def atualizar_log(self, mensagem, tipo="INFO"):
        """Atualiza o log na interface"""
        self.root.after(0, self._adicionar_log, mensagem, tipo)
    
    def _adicionar_log(self, mensagem, tipo):
        """Adiciona log no text widget"""
        self.text_logs.insert(tk.END, mensagem + "\n")
        
        # Colorir por tipo
        if tipo == "ERROR":
            self.text_logs.tag_add("error", f"end-{len(mensagem)+2}c", "end-1c")
            self.text_logs.tag_config("error", foreground="#e74c3c")
        elif tipo == "SUCCESS":
            self.text_logs.tag_add("success", f"end-{len(mensagem)+2}c", "end-1c")
            self.text_logs.tag_config("success", foreground="#27ae60")
        elif tipo == "WARNING":
            self.text_logs.tag_add("warning", f"end-{len(mensagem)+2}c", "end-1c")
            self.text_logs.tag_config("warning", foreground="#f39c12")
        
        self.text_logs.see(tk.END)
    
    def atualizar_estatisticas(self, sucessos, erros):
        """Atualiza as estatísticas"""
        self.sucessos = sucessos
        self.erros = erros
        self.root.after(0, self._atualizar_stats_ui)
    
    def _atualizar_stats_ui(self):
        """Atualiza UI das estatísticas"""
        self.label_sucessos.config(text=str(self.sucessos))
        self.label_erros.config(text=str(self.erros))
    
    def selecionar_planilha(self):
        """Seleciona a planilha"""
        arquivo = filedialog.askopenfilename(
            title="Selecionar Planilha",
            filetypes=[("Planilhas", "*.xlsx *.xls *.csv"), ("Excel", "*.xlsx *.xls"), ("CSV", "*.csv")]
        )
        
        if arquivo:
            self.planilha_path = arquivo
            self.label_planilha.config(text=os.path.basename(arquivo), fg="black")
            
            # Ler planilha
            try:
                caminho = Path(arquivo)
                if caminho.suffix.lower() == '.csv':
                    df = pd.read_csv(caminho, encoding='utf-8')
                else:
                    df = pd.read_excel(caminho, engine='openpyxl')
                
                # Procurar coluna de autos
                coluna_auto = None
                for col in df.columns:
                    if 'auto' in str(col).lower() and 'infração' in str(col).lower():
                        coluna_auto = col
                        break
                
                if coluna_auto is None:
                    coluna_auto = df.columns[0]
                
                self.autos = df[coluna_auto].dropna().astype(str).tolist()
                self.autos = [a for a in self.autos if a.lower() != 'auto de infração' and a.strip() != '']

                # ── Melhoria 4: verificar duplicatas ──
                from collections import Counter as _Counter
                contagem = _Counter(self.autos)
                duplicatas = {a: n for a, n in contagem.items() if n > 1}
                msg_carregado = f"Planilha carregada!\n{len(self.autos)} autos encontrados."
                if duplicatas:
                    lista_dup = "\n".join(f"  {a} ({n}x)" for a, n in list(duplicatas.items())[:10])
                    sufixo = f"\n...e mais {len(duplicatas)-10} outros." if len(duplicatas) > 10 else ""
                    msg_carregado += (
                        f"\n\n⚠️ {len(duplicatas)} auto(s) duplicado(s) detectado(s):\n"
                        f"{lista_dup}{sufixo}\n\n"
                        "Os duplicatas serão processados quantas vezes aparecerem na planilha.\n"
                        "Deseja remover os duplicatas automaticamente?"
                    )
                    remover = messagebox.askyesno("Duplicatas encontradas", msg_carregado)
                    if remover:
                        vistos = set()
                        autos_sem_dup = []
                        for a in self.autos:
                            if a not in vistos:
                                vistos.add(a)
                                autos_sem_dup.append(a)
                        removidos = len(self.autos) - len(autos_sem_dup)
                        self.autos = autos_sem_dup
                        messagebox.showinfo("Duplicatas removidas", f"{removidos} auto(s) duplicado(s) removido(s).\nTotal final: {len(self.autos)} autos.")
                    else:
                        messagebox.showinfo("Planilha carregada", f"Planilha carregada com {len(self.autos)} autos (incluindo duplicatas).")
                else:
                    messagebox.showinfo("Sucesso", msg_carregado)
            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao ler planilha:\n{str(e)}")
    
    def iniciar_automacao(self):
        """Inicia a automação"""
        # Validar
        if not self.logado:
            messagebox.showerror("Erro", "Você precisa fazer login primeiro.")
            return
        
        if not self.planilha_path or not self.autos:
            messagebox.showerror("Erro", "Selecione uma planilha válida.")
            return
        
        # Desabilitar botão iniciar
        self.btn_iniciar.config(state=tk.DISABLED)
        self.btn_pausar.config(state=tk.NORMAL)
        self.btn_parar.config(state=tk.NORMAL)
        self.progress_bar['value'] = 0
        
        # Limpar logs e estatísticas
        self.text_logs.delete(1.0, tk.END)
        self.sucessos = 0
        self.erros = 0
        self.ultimos_autos_com_erro = []
        self.btn_reprocessar_erros.config(state=tk.DISABLED)
        self.btn_gerar_planilha.config(state=tk.DISABLED)
        self._atualizar_stats_ui()

        # Ativar log em arquivo (não perde logs ao fechar a janela)
        try:
            base_dir = Path(__file__).resolve().parent
            log_dir = base_dir / "logs"
            log_dir.mkdir(exist_ok=True)
            nome_arquivo = f"automacao_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
            self.logger.set_log_file(str(log_dir / nome_arquivo))
            self.logger.log(f"Log desta execução: {log_dir / nome_arquivo}", "INFO")
        except Exception:
            self.logger.set_log_file(None)
        
        # Criar automação (reutilizar se já tiver navegador aberto do login)
        tipo = self.tipo_automacao.get()
        if tipo == "consulta":
            self.automacao = AutomacaoConsultaPagamento(self.logger)
        else:
            self.automacao = AutomacaoInscricaoSerasa(self.logger)
        
        # Não reutilizar navegador headless - será criado visível na automação
        
        # Executar em thread
        # ── Melhoria 4: rastreamento de tempo para ETA ──
        self._tempo_inicio_auto = None       # timestamp do início do primeiro auto
        self._idx_atual_eta = 0              # índice processado mais recente
        self._total_autos_eta = len(self.autos)

        self.thread_automacao = threading.Thread(target=self._executar_automacao, 
                                                  args=(self.usuario_logado, self.senha_logada), daemon=True)
        self.thread_automacao.start()
    
    def _executar_automacao(self, usuario, senha):
        """Executa a automação em thread separada"""
        try:
            # ── Melhoria 1: verificar checkpoint ──
            checkpoint = CheckpointManager(self.planilha_path) if self.planilha_path else None
            idx_inicio = 0
            if checkpoint and checkpoint.existe():
                dados_cp = checkpoint.carregar()
                idx_cp = dados_cp.get("idx", 0)
                total_cp = dados_cp.get("total", len(self.autos))
                if 0 < idx_cp < total_cp:
                    resposta = messagebox.askyesno(
                        "Checkpoint encontrado",
                        f"Foi encontrado um progresso anterior: {idx_cp}/{total_cp} autos processados.\n\n"
                        f"Deseja continuar de onde parou (a partir do auto {idx_cp + 1})?\n\n"
                        f"Clique 'Não' para recomeçar do zero."
                    )
                    if resposta:
                        idx_inicio = idx_cp
                        self.logger.log(f"Retomando de onde parou: auto {idx_inicio + 1}/{total_cp}", "INFO")
                    else:
                        checkpoint.limpar()

            # Criar driver visível (não headless) para a automação
            if not self.automacao.criar_driver(headless=False):
                self.root.after(0, self._finalizar_automacao, False)
                return
            
            # Fazer login no navegador visível
            if not self.automacao.fazer_login(usuario, senha):
                self.automacao.fechar()
                self.root.after(0, self._finalizar_automacao, False)
                return
            
            # Navegar
            if not self.automacao.navegar_para_formulario():
                self.automacao.fechar()
                self.root.after(0, self._finalizar_automacao, False)
                return
            
            # Processar autos
            sucessos, erros = self.automacao.processar_autos(
                self.autos,
                progress_callback=self._atualizar_progresso_com_eta,
                stats_callback=self.atualizar_estatisticas,
                error_handler=self._tratar_erro,
                checkpoint=checkpoint,
                idx_inicio=idx_inicio
            )
            
            # Salvar resultados
            caminho_resultado = self.automacao.salvar_resultados(self.planilha_path)

            # ── Melhoria 3: exportar log para aba extra no Excel ──
            if caminho_resultado:
                self.automacao.salvar_log_excel(caminho_resultado)
            
            # Verificar se é automação de Inscrição SERASA (não fechar navegador)
            is_inscricao_serasa = isinstance(self.automacao, AutomacaoInscricaoSerasa)
            
            if is_inscricao_serasa:
                # Para Inscrição SERASA, manter navegador aberto para usuário clicar manualmente
                self.logger.log("Processamento concluído! Navegador permanece aberto para você clicar em 'Incluir na SERASA' manualmente.", "INFO")
            else:
                # Para outras automações, fechar navegador normalmente
                self.automacao.fechar()
            
            # Finalizar
            mensagem = f"Processamento concluído!\n\nSucessos: {sucessos}\nErros: {erros}\n\n"
            if caminho_resultado:
                mensagem += f"Resultados salvos em:\n{caminho_resultado}"
            
            if is_inscricao_serasa:
                mensagem += "\n\nNavegador permanece aberto.\nVocê pode clicar em 'Incluir na SERASA' manualmente quando quiser."
            
            self.root.after(0, self._finalizar_automacao, True, mensagem)
            
        except Exception as e:
            self.logger.log(f"Erro fatal: {str(e)}", "ERROR")
            if self.automacao:
                self.automacao.fechar()
            self.root.after(0, self._finalizar_automacao, False, f"Erro: {str(e)}")
    
    def _atualizar_progresso_com_eta(self, mensagem: str):
        """Atualiza barra determinística, ETA e velocidade em tempo real."""
        import re as _re
        pct = self.progress_bar['value']
        match = _re.search(r"(\d+)/(\d+)", mensagem)
        if match:
            idx = int(match.group(1))
            total = int(match.group(2))
            if total > 0:
                pct = round((idx / total) * 100, 1)
            agora = time.monotonic()
            if self._tempo_inicio_auto is None or idx <= 1:
                self._tempo_inicio_auto = agora
                self._idx_atual_eta = idx
            else:
                concluidos = idx - 1
                if concluidos > 0:
                    tempo_decorrido = agora - self._tempo_inicio_auto
                    media_por_auto = tempo_decorrido / concluidos
                    # ETA
                    restantes = total - idx
                    eta_segundos = int(media_por_auto * restantes)
                    if eta_segundos >= 60:
                        eta_str = f"~{eta_segundos // 60}min {eta_segundos % 60}s restantes"
                    else:
                        eta_str = f"~{eta_segundos}s restantes"
                    # Velocidade (autos/min)
                    vel = round(60 / media_por_auto, 1) if media_por_auto > 0 else 0
                    mensagem = f"{mensagem}  |  {eta_str}  |  {vel} autos/min"
        def _update(m=mensagem, v=pct):
            self.progress_var.set(m)
            self.progress_bar['value'] = v
        self.root.after(0, _update)

    def _atualizar_progresso(self, mensagem):
        """Atualiza mensagem de progresso (versão simples sem ETA)"""
        self.root.after(0, lambda: self.progress_var.set(mensagem))
    
    def _tratar_erro(self, auto, erro):
        """Trata erro durante processamento"""
        # Por padrão, continuar automaticamente
        # O usuário pode pausar manualmente usando o botão
        self.logger.log(f"Erro no auto {auto}: {erro}", "ERROR")
        return "continuar"
    
    def _finalizar_automacao(self, sucesso, mensagem=None):
        """Finaliza a automação"""
        self.progress_bar['value'] = 100 if sucesso else self.progress_bar['value']
        self.btn_iniciar.config(state=tk.NORMAL)
        self.btn_pausar.config(state=tk.DISABLED)
        self.btn_continuar.config(state=tk.DISABLED)
        self.btn_parar.config(state=tk.DISABLED)
        
        # Se foi Inscrição SERASA e há erros, habilitar "Reprocessar apenas erros"
        if sucesso and self.automacao and isinstance(self.automacao, AutomacaoInscricaoSerasa):
            autos_com_erro = [r['auto'] for r in self.automacao.resultados if r.get('situacao') != 'SELECIONADO']
            self.ultimos_autos_com_erro = autos_com_erro
            if autos_com_erro:
                self.btn_reprocessar_erros.config(state=tk.NORMAL)
                self.logger.log(f"Você pode clicar em 'Reprocessar apenas erros' para rodar de novo os {len(autos_com_erro)} auto(s) que falharam.", "INFO")
            else:
                self.btn_reprocessar_erros.config(state=tk.DISABLED)
        else:
            self.btn_reprocessar_erros.config(state=tk.DISABLED)

        # Habilitar "Gerar Planilha de Resultado" se houver resultados disponíveis
        if sucesso and self.automacao and self.automacao.resultados and self.planilha_path:
            self.btn_gerar_planilha.config(state=tk.NORMAL)
        else:
            self.btn_gerar_planilha.config(state=tk.DISABLED)
        
        # ── Melhoria 2: resumo detalhado por status ──
        if sucesso and self.automacao and self.automacao.resultados:
            from collections import Counter as _Counter
            contagem = _Counter(r.get('situacao', 'DESCONHECIDO') for r in self.automacao.resultados)
            linhas_resumo = [f"  {status}: {qtd}" for status, qtd in sorted(contagem.items())]
            resumo = "\n".join(linhas_resumo)
            if mensagem:
                mensagem = mensagem.rstrip() + f"\n\n── Resumo por status ──\n{resumo}"

        if mensagem:
            if sucesso:
                messagebox.showinfo("Concluído", mensagem)
            else:
                messagebox.showerror("Erro", mensagem)
    
    def pausar_automacao(self):
        """Pausa a automação"""
        if self.automacao:
            self.automacao.pausado = True
            self.btn_pausar.config(state=tk.DISABLED)
            self.btn_continuar.config(state=tk.NORMAL)
            self.logger.log("Automação pausada", "WARNING")
    
    def continuar_automacao(self):
        """Continua a automação"""
        if self.automacao:
            self.automacao.pausado = False
            self.btn_pausar.config(state=tk.NORMAL)
            self.btn_continuar.config(state=tk.DISABLED)
            self.logger.log("Automação continuada", "INFO")
    
    def parar_automacao(self):
        """Para a automação"""
        if self.automacao:
            self.automacao.parar = True
            self.automacao.pausado = False
            self.logger.log("Automação interrompida pelo usuário", "WARNING")
            self._finalizar_automacao(False, "Automação interrompida.")
    
    def reprocessar_apenas_erros(self):
        """Roda de novo apenas os autos que deram erro na última execução (para confirmar que foram erros de fato)."""
        if not self.ultimos_autos_com_erro:
            messagebox.showinfo("Reprocessar erros", "Não há autos com erro da última execução.")
            return
        if not self.automacao or not isinstance(self.automacao, AutomacaoInscricaoSerasa):
            messagebox.showwarning("Reprocessar erros", "Navegador da Inscrição SERASA não está disponível. Rode uma automação de Inscrição primeiro.")
            return
        if not self.planilha_path:
            messagebox.showwarning("Reprocessar erros", "Planilha original não definida.")
            return
        n = len(self.ultimos_autos_com_erro)
        if not messagebox.askyesno("Reprocessar apenas erros", 
            f"Rodar novamente apenas os {n} auto(s) que deram erro?\n\nIsso usa o mesmo navegador já aberto. O resultado será salvo em um arquivo separado (Reprocessamento)."):
            return
        self.btn_reprocessar_erros.config(state=tk.DISABLED)
        self.btn_gerar_planilha.config(state=tk.DISABLED)
        self.btn_iniciar.config(state=tk.DISABLED)
        self.btn_pausar.config(state=tk.NORMAL)
        self.btn_parar.config(state=tk.NORMAL)
        self.progress_bar['value'] = 0
        self.text_logs.insert(tk.END, "\n--- Reprocessando apenas erros ---\n", "INFO")
        self.sucessos = 0
        self.erros = 0
        self._atualizar_stats_ui()
        # Resetar variáveis de ETA para o reprocessamento
        self._tempo_inicio_auto = None
        self._idx_atual_eta = 0
        self._total_autos_eta = len(self.ultimos_autos_com_erro)
        self.thread_automacao = threading.Thread(target=self._executar_reprocessar_erros, daemon=True)
        self.thread_automacao.start()
    
    def _executar_reprocessar_erros(self):
        """Executa o reprocessamento apenas dos autos que deram erro (mesmo navegador)."""
        try:
            autos_erro = list(self.ultimos_autos_com_erro)
            self.automacao.resultados = []
            self.automacao.parar = False
            self.automacao.pausado = False
            sucessos, erros = self.automacao.processar_autos(
                autos_erro,
                progress_callback=self._atualizar_progresso_com_eta,
                stats_callback=self.atualizar_estatisticas,
                error_handler=self._tratar_erro
            )
            caminho_resultado = self.automacao.salvar_resultados(self.planilha_path, sufixo_arquivo="Reprocessamento")
            if caminho_resultado:
                self.automacao.salvar_log_excel(caminho_resultado)
            self.logger.log("Reprocessamento concluído! Navegador permanece aberto.", "INFO")
            msg = f"Reprocessamento concluído!\n\nSucessos: {sucessos}\nErros: {erros}\n\n"
            if caminho_resultado:
                msg += f"Resultados salvos em:\n{caminho_resultado}"
            self.root.after(0, self._finalizar_automacao, True, msg)
        except Exception as e:
            self.logger.log(f"Erro no reprocessamento: {str(e)}", "ERROR")
            self.root.after(0, self._finalizar_automacao, False, f"Erro: {str(e)}")
    
    def gerar_planilha_resultado(self):
        """Gera/regenera a planilha de resultado com o timestamp atual."""
        if not self.automacao or not self.automacao.resultados:
            messagebox.showwarning("Gerar Planilha", "Não há resultados disponíveis para gerar a planilha.\nExecute uma automação primeiro.")
            return
        if not self.planilha_path:
            messagebox.showwarning("Gerar Planilha", "Planilha original não definida.")
            return
        try:
            caminho_resultado = self.automacao.salvar_resultados(self.planilha_path)
            if caminho_resultado:
                self.logger.log(f"Planilha de resultado gerada: {caminho_resultado}", "SUCCESS")
                messagebox.showinfo("Planilha Gerada", f"Planilha gerada com sucesso!\n\n{caminho_resultado}")
            else:
                messagebox.showerror("Erro", "Não foi possível gerar a planilha.")
        except Exception as e:
            self.logger.log(f"Erro ao gerar planilha: {str(e)}", "ERROR")
            messagebox.showerror("Erro", f"Erro ao gerar planilha:\n{str(e)}")

    def exportar_log_agora(self):
        """Exporta o log atual para um arquivo .txt imediatamente."""
        try:
            base_dir = Path(__file__).resolve().parent
            log_dir = base_dir / "logs"
            log_dir.mkdir(exist_ok=True)
            nome_arquivo = f"log_exportado_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
            caminho = log_dir / nome_arquivo
            conteudo = self.logger.get_logs()
            if not conteudo.strip():
                messagebox.showinfo("Exportar Log", "Não há logs para exportar ainda.")
                return
            caminho.write_text(conteudo, encoding="utf-8")
            self.logger.log(f"Log exportado manualmente: {caminho}", "INFO")
            messagebox.showinfo("Log exportado", f"Log salvo com sucesso em:\n{caminho}")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao exportar log:\n{str(e)}")

    def executar(self):
        """Executa a interface"""
        self.root.mainloop()


if __name__ == "__main__":
    app = InterfaceGrafica()
    app.executar()

